# -*- coding: utf-8 -*-
"""
Read process.xlsx `curves`, level at ~1000Hz, propose outliers (high-Q exempt),
then after --exclude confirmation rank by tightest ±δ vs mean_after.
"""
from __future__ import annotations

import argparse
import math
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Same directory import
sys.path.insert(0, str(Path(__file__).resolve().parent))
from outliers import propose_outliers  # noqa: E402
from params_io import (  # noqa: E402
    load_params,
    patch_process_md_section,
    resolve_under_output,
)


def _read_curves(ws) -> tuple[list[float], list[str], list[list[float | None]]]:
    samples: list[str] = []
    col = 2
    while True:
        name = ws.cell(1, col).value
        if name is None or str(name).strip() == "":
            break
        samples.append(str(name).strip())
        col += 1
    if not samples:
        raise SystemExit("no sample names on row 1 starting at column B")

    freqs: list[float] = []
    columns: list[list[float | None]] = [[] for _ in samples]
    row = 2
    while True:
        f = ws.cell(row, 1).value
        if f is None:
            break
        freqs.append(float(f))
        for i in range(len(samples)):
            raw = ws.cell(row, i + 2).value
            columns[i].append(float(raw) if raw is not None and raw != "" else None)
        row += 1
    if not freqs:
        raise SystemExit("no frequency rows in curves")
    return freqs, samples, columns


def _nearest_1000(
    freqs: list[float], vals: list[float | None]
) -> tuple[float, float]:
    pairs = [(f, v) for f, v in zip(freqs, vals) if v is not None]
    if not pairs:
        raise ValueError("empty column")
    f, v = min(pairs, key=lambda p: abs(p[0] - 1000.0))
    return float(f), float(v)


def _level_column(
    freqs: list[float], vals: list[float | None]
) -> tuple[list[float | None], float, float]:
    ref_f, ref_v = _nearest_1000(freqs, vals)
    leveled = [None if v is None else (v - ref_v) for v in vals]
    return leveled, ref_f, ref_v


def _mean_curve(columns: list[list[float | None]]) -> list[float | None]:
    n_rows = len(columns[0])
    out: list[float | None] = []
    for r in range(n_rows):
        vals = [c[r] for c in columns if c[r] is not None]
        out.append(sum(vals) / len(vals) if vals else None)
    return out


def _tightest_delta(
    freqs: list[float],
    sample: list[float | None],
    mean: list[float | None],
    f_lo: float,
    f_hi: float,
    step: float = 0.1,
) -> tuple[float, float, float, int, float | None]:
    """Return (delta, rms, max_abs, n_points, freq_at_max_abs)."""
    devs: list[tuple[float, float]] = []
    for f, s, m in zip(freqs, sample, mean):
        if s is None or m is None:
            continue
        if f < f_lo or f > f_hi:
            continue
        devs.append((f, s - m))
    if not devs:
        raise ValueError("no points in analysis band")
    max_abs = max(abs(d) for _, d in devs)
    candidates = [f for f, d in devs if abs(abs(d) - max_abs) < 1e-9]
    freq_at = min(candidates) if candidates else None
    delta = max(step, math.ceil(max_abs / step - 1e-12) * step)
    delta = round(delta, 10)
    rms = math.sqrt(sum(d * d for _, d in devs) / len(devs))
    return delta, rms, max_abs, len(devs), freq_at


def _replace_sheet(wb, name: str) -> Worksheet:
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def _write_wide(
    ws: Worksheet,
    freqs: list[float],
    samples: list[str],
    columns: list[list[float | None]],
    freq_header: str = "freq_hz",
) -> None:
    ws.append([freq_header, *samples])
    for r, f in enumerate(freqs):
        ws.append([f, *[columns[i][r] for i in range(len(samples))]])


def _write_mean_sheet(
    wb, name: str, freqs: list[float], mean: list[float | None], leveled: list[list]
) -> None:
    ws = _replace_sheet(wb, name)
    ws.append(["freq_hz", "mean_leveled_db", "n_samples_at_freq"])
    for r, f in enumerate(freqs):
        vals = [c[r] for c in leveled if c[r] is not None]
        ws.append([f, mean[r], len(vals)])


def _parse_exclude(raw: str | None) -> tuple[bool, list[str]]:
    """
    Returns (confirmed, exclude_list).
    confirmed=False when exclude omitted / null (gate pending).
    """
    if raw is None:
        return False, []
    s = raw.strip()
    if s.lower() == "none" or s == "":
        return True, []
    names = [p.strip() for p in s.split(",") if p.strip()]
    return True, names


def _exclude_from_params(curves: dict) -> str | None:
    """
    params.curves.exclude:
      null / missing -> pending (None)
      [] / "none" -> confirmed none
      ["A","B"] or "A,B" -> confirmed list
    """
    if "exclude" not in curves:
        return None
    raw = curves.get("exclude")
    if raw is None:
        return None
    if isinstance(raw, list):
        if not raw:
            return "none"
        return ",".join(str(x).strip() for x in raw if str(x).strip())
    if isinstance(raw, str):
        return raw
    raise SystemExit("curves.exclude must be null, list, or string")


def _fmt_list(names: list[str]) -> str:
    return "、".join(names) if names else "无"


def _write_outlier_pending_md(
    process_md: Path,
    *,
    f_lo: float,
    f_hi: float,
    q_threshold: float,
    suggested: list[str],
    exempted: list[str],
    review_rows: list[dict],
) -> None:
    table_rows = []
    for row in review_rows:
        table_rows.append(
            "| {sample} | {max_abs_dev_db} | {freq_at_max_hz} | {q_estimate} | "
            "{q_stable} | {exempt} | {suggest} | {reason} |".format(
                sample=row["sample"],
                max_abs_dev_db=row["max_abs_dev_db"],
                freq_at_max_hz=row["freq_at_max_hz"],
                q_estimate=row["q_estimate"],
                q_stable="是" if row["q_stable"] else "否",
                exempt="是" if row["exempt_high_q"] else "否",
                suggest="是" if row["suggest_exclude"] else "否",
                reason=row["suggest_reason"],
            )
        )
    body = "\n".join(
        [
            "## 奇异值分析",
            "",
            f"- 分析频段：{f_lo}–{f_hi} Hz",
            f"- Q 豁免阈值：Q ≥ {q_threshold}",
            "- 提名门槛：max|dev| ≥ max(1.0 dB, 2×批次中位数)",
            "- 门禁状态：pending（待用户确认）",
            f"- 脚本建议剔除：{_fmt_list(suggested)}",
            f"- 高Q豁免（不建议剔除）：{_fmt_list(exempted)}",
            "",
            "### 提名明细（来自 outlier_review）",
            "",
            "| 样机 | maxAbsDev | 频率Hz | Q | Q稳定 | 高Q豁免 | 建议剔除 | 理由 |",
            "|------|-----------|--------|---|-------|---------|----------|------|",
            *table_rows,
            "",
            "### 用户决定",
            "",
            "待确认。请回复：不剔除 / 按建议剔除 / 自定义名单（样机名逗号分隔）。",
            "",
        ]
    )
    patch_process_md_section(process_md, "奇异值分析", body)


def _write_outlier_confirmed_and_curves_md(
    process_md: Path,
    *,
    f_lo: float,
    f_hi: float,
    q_threshold: float,
    suggested: list[str],
    exempted: list[str],
    review_rows: list[dict],
    exclude_list: list[str],
    n_kept: int,
    golden: str,
    aux: list[str],
    rank_rows: list[tuple],
) -> None:
    table_rows = []
    for row in review_rows:
        table_rows.append(
            "| {sample} | {max_abs_dev_db} | {freq_at_max_hz} | {q_estimate} | "
            "{q_stable} | {exempt} | {suggest} | {reason} |".format(
                sample=row["sample"],
                max_abs_dev_db=row["max_abs_dev_db"],
                freq_at_max_hz=row["freq_at_max_hz"],
                q_estimate=row["q_estimate"],
                q_stable="是" if row["q_stable"] else "否",
                exempt="是" if row["exempt_high_q"] else "否",
                suggest="是" if row["suggest_exclude"] else "否",
                reason=row["suggest_reason"],
            )
        )
    decision = "不剔除" if not exclude_list else f"剔除 {_fmt_list(exclude_list)}"
    outlier_body = "\n".join(
        [
            "## 奇异值分析",
            "",
            f"- 分析频段：{f_lo}–{f_hi} Hz",
            f"- Q 豁免阈值：Q ≥ {q_threshold}",
            "- 提名门槛：max|dev| ≥ max(1.0 dB, 2×批次中位数)",
            "- 门禁状态：confirmed",
            f"- 脚本建议剔除：{_fmt_list(suggested)}",
            f"- 高Q豁免（不建议剔除）：{_fmt_list(exempted)}",
            "",
            "### 提名明细（来自 outlier_review）",
            "",
            "| 样机 | maxAbsDev | 频率Hz | Q | Q稳定 | 高Q豁免 | 建议剔除 | 理由 |",
            "|------|-----------|--------|---|-------|---------|----------|------|",
            *table_rows,
            "",
            "### 用户决定",
            "",
            f"- 确认结果：{decision}",
            f"- 进入 mean_after 的样机数：{n_kept}",
            "- 详见 `outlier_decision` sheet",
            "",
        ]
    )
    patch_process_md_section(process_md, "奇异值分析", outlier_body)

    rank_md = []
    for i, (name, delta, rms, *_rest) in enumerate(rank_rows, start=1):
        if i == 1:
            role = "golden"
        elif name in aux:
            role = "aux"
        else:
            role = "—"
        rank_md.append(f"| {name} | {delta} | {rms} | {i} | {role} |")
    curves_body = "\n".join(
        [
            "## 曲线分析",
            "",
            f"- 金标：{golden}",
            f"- 辅标：{_fmt_list(aux)}",
            f"- 分析频段：{f_lo}–{f_hi} Hz",
            f"- 均值：剔异后（mean_after）；剔除名单：{_fmt_list(exclude_list)}",
            "",
            "| 样机 | 最紧δ | RMS | 排名 | 角色 |",
            "|------|-------|-----|------|------|",
            *rank_md,
            "",
        ]
    )
    patch_process_md_section(process_md, "曲线分析", curves_body)


def run(
    xlsx: Path,
    aux_count: int,
    f_lo: float,
    f_hi: float,
    exclude_arg: str | None,
    q_threshold: float = 10.0,
    process_md: Path | None = None,
) -> None:
    wb = load_workbook(xlsx)
    if "curves" not in wb.sheetnames:
        raise SystemExit(f"missing sheet 'curves' in {xlsx}")
    freqs, samples, columns = _read_curves(wb["curves"])

    # --- step1: 1kHz level refs + leveled curves ---
    leveled: list[list[float | None]] = []
    refs: list[tuple[str, float, float]] = []
    for name, col in zip(samples, columns):
        lev, ref_f, ref_v = _level_column(freqs, col)
        leveled.append(lev)
        refs.append((name, ref_f, ref_v))

    ref_ws = _replace_sheet(wb, "curve_step1_level_ref")
    ref_ws.append(
        [
            "sample",
            "target_hz",
            "ref_freq_hz",
            "abs_freq_error_hz",
            "ref_level_db",
            "level_formula",
        ]
    )
    for name, ref_f, ref_v in refs:
        ref_ws.append(
            [
                name,
                1000.0,
                ref_f,
                abs(ref_f - 1000.0),
                ref_v,
                "leveled = raw - ref_level_db",
            ]
        )

    lev_ws = _replace_sheet(wb, "curve_step1_leveled")
    _write_wide(lev_ws, freqs, samples, leveled)

    # --- mean_before (full set) ---
    mean_before = _mean_curve(leveled)
    _write_mean_sheet(wb, "curve_mean_before_outlier", freqs, mean_before, leveled)
    # keep legacy alias for older readers
    _write_mean_sheet(wb, "curve_step2_mean", freqs, mean_before, leveled)

    # --- outlier proposal ---
    mean_for_q = [0.0 if m is None else float(m) for m in mean_before]
    sample_map: dict[str, list[float]] = {}
    for name, col in zip(samples, leveled):
        sample_map[name] = [0.0 if v is None else float(v) for v in col]

    review_rows = propose_outliers(
        freqs,
        sample_map,
        mean_for_q,
        f_lo=f_lo,
        f_hi=f_hi,
        q_threshold=q_threshold,
    )

    rev = _replace_sheet(wb, "outlier_review")
    rev.append(
        [
            "sample",
            "max_abs_dev_db",
            "freq_at_max_hz",
            "q_estimate",
            "q_stable",
            "exempt_high_q",
            "exempt_reason",
            "median_max_abs_db",
            "threshold_db",
            "suggest_exclude",
            "suggest_reason",
            "f_lo_hz",
            "f_hi_hz",
            "q_threshold",
        ]
    )
    for row in review_rows:
        rev.append(
            [
                row["sample"],
                row["max_abs_dev_db"],
                row["freq_at_max_hz"],
                row["q_estimate"],
                "yes" if row["q_stable"] else "no",
                "yes" if row["exempt_high_q"] else "no",
                row["exempt_reason"],
                row.get("median_max_abs_db"),
                row.get("threshold_db"),
                "yes" if row["suggest_exclude"] else "no",
                row["suggest_reason"],
                f_lo,
                f_hi,
                q_threshold,
            ]
        )

    suggested = [r["sample"] for r in review_rows if r["suggest_exclude"]]
    exempted = [r["sample"] for r in review_rows if r["exempt_high_q"]]

    confirmed, exclude_list = _parse_exclude(exclude_arg)

    meta = _replace_sheet(wb, "curve_meta")
    meta.append(["key", "value"])
    meta.append(["aux_count", aux_count])
    meta.append(["f_lo_hz", f_lo])
    meta.append(["f_hi_hz", f_hi])
    meta.append(["delta_step_db", 0.1])
    meta.append(["n_samples_all", len(samples)])
    meta.append(["q_threshold", q_threshold])
    meta.append(["outlier_propose_min_dev_db", 1.0])
    meta.append(["outlier_propose_median_factor", 2.0])
    meta.append(["suggested_exclude", ", ".join(suggested) if suggested else "(none)"])
    meta.append(["exempt_high_q", ", ".join(exempted) if exempted else "(none)"])

    if not confirmed:
        meta.append(["outlier_gate", "pending"])
        meta.append(["excluded", "(pending user confirm)"])
        meta.append(
            [
                "process_sheets",
                "curve_step1_level_ref | curve_step1_leveled | "
                "curve_mean_before_outlier | outlier_review | curve_meta "
                "(rank sheets after --exclude)",
            ]
        )
        # Remove stale rank sheets if re-running propose
        for stale in (
            "outlier_decision",
            "curve_mean_after_outlier",
            "curve_step3_dev",
            "curve_step3_dev_in_band",
            "curve_step4_envelope",
            "curve_rank",
        ):
            if stale in wb.sheetnames:
                del wb[stale]
        wb.save(xlsx)
        if process_md is not None:
            _write_outlier_pending_md(
                process_md,
                f_lo=f_lo,
                f_hi=f_hi,
                q_threshold=q_threshold,
                suggested=suggested,
                exempted=exempted,
                review_rows=review_rows,
            )
            print(f"OK patched process.md section 奇异值分析 -> {process_md}")
        print(f"OK wrote outlier proposal -> {xlsx}")
        print(
            f"outlier_gate=pending suggested={suggested or ['(none)']} "
            f"exempt_high_q={exempted or ['(none)']}"
        )
        print(
            "Confirm then set params.curves.exclude to [] or [names], then re-run"
        )
        return

    # --- confirmed: apply excludes, mean_after, rank ---
    unknown = [n for n in exclude_list if n not in samples]
    if unknown:
        raise SystemExit(f"unknown sample(s) in --exclude: {unknown}")

    exclude_set = set(exclude_list)
    kept_names = [n for n in samples if n not in exclude_set]
    kept_idx = [i for i, n in enumerate(samples) if n not in exclude_set]
    if not kept_names:
        raise SystemExit("all samples excluded; nothing left to rank")

    # decision sheet
    dec = _replace_sheet(wb, "outlier_decision")
    dec.append(
        [
            "sample",
            "suggest_exclude",
            "exempt_high_q",
            "user_exclude",
            "in_mean_after",
            "decision_note",
        ]
    )
    suggest_map = {r["sample"]: r for r in review_rows}
    for name in samples:
        r = suggest_map[name]
        user_ex = name in exclude_set
        note = []
        if r["exempt_high_q"] and user_ex:
            note.append("高Q豁免样机仍被用户剔除")
        if r["suggest_exclude"] and not user_ex:
            note.append("脚本建议剔除但用户保留")
        if not r["suggest_exclude"] and user_ex and not r["exempt_high_q"]:
            note.append("脚本未建议但用户剔除")
        if user_ex:
            note.append("已剔除，不进 mean_after")
        else:
            note.append("保留，进入 mean_after")
        dec.append(
            [
                name,
                "yes" if r["suggest_exclude"] else "no",
                "yes" if r["exempt_high_q"] else "no",
                "yes" if user_ex else "no",
                "no" if user_ex else "yes",
                "；".join(note),
            ]
        )

    leveled_kept = [leveled[i] for i in kept_idx]
    mean_after = _mean_curve(leveled_kept)
    _write_mean_sheet(wb, "curve_mean_after_outlier", freqs, mean_after, leveled_kept)

    # deviation vs mean_after (all samples for traceability; ranking uses kept only)
    dev_cols: list[list[float | None]] = []
    for col in leveled:
        dev_cols.append(
            [
                None if s is None or m is None else (s - m)
                for s, m in zip(col, mean_after)
            ]
        )
    dev_ws = _replace_sheet(wb, "curve_step3_dev")
    _write_wide(dev_ws, freqs, samples, dev_cols)

    band_ws = _replace_sheet(wb, "curve_step3_dev_in_band")
    band_ws.append(["freq_hz", "in_analysis_band", *samples])
    for r, f in enumerate(freqs):
        in_band = "yes" if f_lo <= f <= f_hi else "no"
        row_vals: list[float | None | str] = [f, in_band]
        for i in range(len(samples)):
            row_vals.append(None if in_band == "no" else dev_cols[i][r])
        band_ws.append(row_vals)

    rows: list[tuple[str, float, float, float, int, float | None]] = []
    for name, col in zip(kept_names, leveled_kept):
        delta, rms, max_abs, n_pts, freq_at = _tightest_delta(
            freqs, col, mean_after, f_lo, f_hi
        )
        rows.append((name, delta, rms, max_abs, n_pts, freq_at))
    rows.sort(key=lambda r: (r[1], r[2], r[0]))

    stats = _replace_sheet(wb, "curve_step4_envelope")
    stats.append(
        [
            "sample",
            "max_abs_dev_db",
            "freq_at_max_hz",
            "delta_db",
            "rms_db",
            "n_points_in_band",
            "f_lo_hz",
            "f_hi_hz",
            "delta_step_db",
            "rank",
            "role",
            "excluded_from_mean",
        ]
    )
    for i, (name, delta, rms, max_abs, n_pts, freq_at) in enumerate(rows, start=1):
        if i == 1:
            role = "golden"
        elif i <= 1 + aux_count:
            role = "aux"
        else:
            role = ""
        stats.append(
            [
                name,
                max_abs,
                freq_at,
                delta,
                rms,
                n_pts,
                f_lo,
                f_hi,
                0.1,
                i,
                role,
                "no",
            ]
        )

    rank_ws = _replace_sheet(wb, "curve_rank")
    rank_ws.append(
        [
            "sample",
            "max_abs_dev",
            "freq_at_max_abs",
            "delta",
            "rms",
            "n_points_in_band",
            "rank",
            "role",
        ]
    )
    for i, (name, delta, rms, max_abs, n_pts, freq_at) in enumerate(rows, start=1):
        if i == 1:
            role = "golden"
        elif i <= 1 + aux_count:
            role = "aux"
        else:
            role = ""
        rank_ws.append([name, max_abs, freq_at, delta, rms, n_pts, i, role])

    meta.append(["outlier_gate", "confirmed"])
    meta.append(
        ["excluded", ", ".join(exclude_list) if exclude_list else "(none)"]
    )
    meta.append(["n_samples_after_exclude", len(kept_names)])
    if rows:
        meta.append(["golden", rows[0][0]])
        aux = [r[0] for r in rows[1 : 1 + aux_count]]
        meta.append(["aux", ", ".join(aux)])
    meta.append(
        [
            "process_sheets",
            "curve_step1_level_ref | curve_step1_leveled | "
            "curve_mean_before_outlier | outlier_review | outlier_decision | "
            "curve_mean_after_outlier | curve_step3_dev | curve_step3_dev_in_band | "
            "curve_step4_envelope | curve_rank | curve_meta",
        ]
    )

    wb.save(xlsx)
    golden = rows[0][0] if rows else ""
    aux = [r[0] for r in rows[1 : 1 + aux_count]] if rows else []
    if process_md is not None:
        _write_outlier_confirmed_and_curves_md(
            process_md,
            f_lo=f_lo,
            f_hi=f_hi,
            q_threshold=q_threshold,
            suggested=suggested,
            exempted=exempted,
            review_rows=review_rows,
            exclude_list=exclude_list,
            n_kept=len(kept_names),
            golden=golden,
            aux=aux,
            rank_rows=rows,
        )
        print(
            f"OK patched process.md sections 奇异值分析 + 曲线分析 -> {process_md}"
        )
    print(f"OK wrote curve process sheets -> {xlsx}")
    print(
        f"outlier_gate=confirmed excluded={exclude_list or ['(none)']} "
        f"golden={golden} aux_count={aux_count} n_kept={len(kept_names)}"
    )


def main() -> None:
    p = argparse.ArgumentParser(
        description="Level@1kHz, propose outliers, rank; inputs from params.json"
    )
    p.add_argument(
        "--params",
        type=Path,
        required=True,
        help="Path to params.json (deterministic inputs)",
    )
    args = p.parse_args()
    params = load_params(args.params)
    xlsx = resolve_under_output(params, "process_xlsx", args.params)
    md = resolve_under_output(params, "process_md", args.params)
    if not xlsx.is_file():
        raise SystemExit(f"file not found: {xlsx}")
    if not md.is_file():
        raise SystemExit(f"process.md not found: {md}")

    curves = params.get("curves") or {}
    aux_count = int(curves.get("aux_count", 3))
    f_lo = float(params.get("f_lo_hz", 100))
    f_hi = float(params.get("f_hi_hz", 15000))
    q_threshold = float(curves.get("q_threshold", 10))
    exclude_arg = _exclude_from_params(curves)

    if aux_count < 0:
        raise SystemExit("curves.aux_count must be >= 0")
    run(
        xlsx,
        aux_count,
        f_lo,
        f_hi,
        exclude_arg,
        q_threshold,
        process_md=md,
    )


if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""Tests for run_sensitivity.py: 1000Hz picking, midline modes, 0.5dB binning."""
from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

SCRIPTS = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(SCRIPTS))

from conftest import build_curves_workbook, write_process_md  # noqa: E402
from run_sensitivity import _bin_index, run  # noqa: E402

FREQS = [500.0, 1000.0, 2000.0]


def _make_batch(tmp_path: Path, values_by_sample: dict[str, list[float]]) -> tuple[Path, Path]:
    xlsx = tmp_path / "process.xlsx"
    md = tmp_path / "process.md"
    build_curves_workbook(
        xlsx, samples=list(values_by_sample), freqs=FREQS, values_by_sample=values_by_sample
    )
    write_process_md(md)
    return xlsx, md


def test_bin_index_edges():
    # width=0.5: [0, 0.5) -> 0; [0.5, 1.0) -> 1; [-0.5, 0) -> -1
    assert _bin_index(0.0) == 0
    assert _bin_index(0.49) == 0
    assert _bin_index(0.5) == 1
    assert _bin_index(-0.01) == -1
    assert _bin_index(-0.5) == -1
    assert _bin_index(-0.51) == -2


def test_midline_mode_mean_computes_average_of_1000hz_levels(tmp_path):
    xlsx, md = _make_batch(
        tmp_path,
        {
            "S1": [10.0, 100.0, 10.0],
            "S2": [10.0, 102.0, 10.0],
            "S3": [10.0, 104.0, 10.0],
        },
    )
    run(xlsx, "mean", None, "dB", process_md=md)

    wb = load_workbook(xlsx)
    meta = {row[0].value: row[1].value for row in wb["sensitivity_meta"].iter_rows(min_row=2)}
    assert meta["midline_mode"] == "mean"
    assert meta["midline_db"] == pytest.approx(102.0)
    # golden = sample closest to midline (102.0) -> S2 (exact match)
    assert meta["golden"] == "S2"


def test_midline_mode_value_uses_provided_midline(tmp_path):
    xlsx, md = _make_batch(
        tmp_path,
        {
            "S1": [10.0, 100.0, 10.0],
            "S2": [10.0, 105.0, 10.0],
        },
    )
    run(xlsx, "value", 100.0, "dB", process_md=md)

    wb = load_workbook(xlsx)
    meta = {row[0].value: row[1].value for row in wb["sensitivity_meta"].iter_rows(min_row=2)}
    assert meta["midline_mode"] == "value"
    assert meta["midline_db"] == pytest.approx(100.0)
    assert meta["golden"] == "S1"


def test_midline_mode_value_without_midline_db_raises(tmp_path):
    xlsx, md = _make_batch(tmp_path, {"S1": [10.0, 100.0, 10.0]})
    with pytest.raises(SystemExit):
        run(xlsx, "value", None, "dB", process_md=md)


def test_sensitivity_sheet_bins_written_correctly(tmp_path):
    xlsx, md = _make_batch(
        tmp_path,
        {
            "S1": [10.0, 100.0, 10.0],  # delta 0 vs midline 100 -> bin 0
            "S2": [10.0, 100.6, 10.0],  # delta +0.6 -> bin 1
            "S3": [10.0, 99.4, 10.0],  # delta -0.6 -> bin -2 (floor(-0.6/0.5)=-2)
        },
    )
    run(xlsx, "value", 100.0, "dB", process_md=md)

    wb = load_workbook(xlsx)
    ws = wb["sensitivity"]
    headers = [c.value for c in ws[1]]
    rows = {row[headers.index("sample")]: row for row in ws.iter_rows(min_row=2, values_only=True)}
    assert rows["S1"][headers.index("bin")] == 0
    assert rows["S2"][headers.index("bin")] == 1
    assert rows["S3"][headers.index("bin")] == -2


def test_process_md_patched_with_sensitivity_section(tmp_path):
    xlsx, md = _make_batch(tmp_path, {"S1": [10.0, 100.0, 10.0], "S2": [10.0, 101.0, 10.0]})
    run(xlsx, "mean", None, "dB", process_md=md)
    text = md.read_text(encoding="utf-8")
    assert "## 灵敏度分析" in text
    assert "S1" in text and "S2" in text
    assert "## 奇异值分析" in text  # sibling section untouched/present

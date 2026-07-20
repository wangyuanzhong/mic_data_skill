# -*- coding: utf-8 -*-
"""Tests for run_curves.py: leveling, outlier gate (pending/confirmed), ranking."""
from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

SCRIPTS = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(SCRIPTS))

from conftest import build_curves_workbook, write_process_md  # noqa: E402
from run_curves import run  # noqa: E402

FREQS = [500.0, 1000.0, 2000.0, 4000.0]


def _make_batch(tmp_path: Path, values_by_sample: dict[str, list[float]]) -> tuple[Path, Path]:
    xlsx = tmp_path / "process.xlsx"
    md = tmp_path / "process.md"
    build_curves_workbook(
        xlsx, samples=list(values_by_sample), freqs=FREQS, values_by_sample=values_by_sample
    )
    write_process_md(md)
    return xlsx, md


def _sheet_rows(xlsx: Path, sheet: str) -> list[tuple]:
    wb = load_workbook(xlsx)
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    return headers, list(ws.iter_rows(min_row=2, values_only=True))


def test_pending_gate_when_exclude_is_none(tmp_path):
    """exclude_arg=None means params.curves.exclude was null -> gate stays pending."""
    xlsx, md = _make_batch(
        tmp_path,
        {
            "S1": [0.0, 0.0, 0.0, 0.0],
            "S2": [0.1, 0.0, -0.1, 0.0],
            "S3": [0.2, 0.0, -0.2, 0.0],
        },
    )
    run(xlsx, aux_count=3, f_lo=100.0, f_hi=15000.0, exclude_arg=None, process_md=md)

    wb = load_workbook(xlsx)
    assert "curve_rank" not in wb.sheetnames
    assert "outlier_decision" not in wb.sheetnames
    meta = {row[0].value: row[1].value for row in wb["curve_meta"].iter_rows(min_row=2)}
    assert meta["outlier_gate"] == "pending"

    text = md.read_text(encoding="utf-8")
    assert "门禁状态：pending" in text


def _expected_rank_order(
    values_by_sample: dict[str, list[float]], freqs: list[float], f_lo: float, f_hi: float
) -> list[str]:
    """Independently recompute the golden-analysis ranking spec for assertion purposes:
    level each curve at its own nearest-1000Hz value, average the leveled curves to
    get mean_after, then rank samples by smallest max|leveled - mean_after| in-band."""
    samples = list(values_by_sample)

    def nearest_1000(vals: list[float]) -> float:
        idx = min(range(len(freqs)), key=lambda i: abs(freqs[i] - 1000.0))
        return vals[idx]

    leveled = {s: [v - nearest_1000(values_by_sample[s]) for v in values_by_sample[s]] for s in samples}
    mean_after = [sum(leveled[s][i] for s in samples) / len(samples) for i in range(len(freqs))]

    def max_abs_dev(sample: str) -> float:
        devs = [
            abs(leveled[sample][i] - mean_after[i])
            for i, f in enumerate(freqs)
            if f_lo <= f <= f_hi
        ]
        return max(devs)

    return sorted(samples, key=lambda s: (max_abs_dev(s), s))


def test_confirmed_gate_ranks_by_tightest_delta(tmp_path):
    """exclude_arg='none' means confirmed with nothing excluded; ranking by tightest
    delta-to-mean_after, matching an independently recomputed expected order."""
    values = {
        "S1": [0.1, 0.0, -0.1, 0.05],
        "S2": [1.0, 0.0, -1.0, 0.0],
        "S3": [-2.0, 0.0, 2.0, -0.3],
    }
    xlsx, md = _make_batch(tmp_path, values)
    run(xlsx, aux_count=1, f_lo=100.0, f_hi=15000.0, exclude_arg="none", process_md=md)

    expected_order = _expected_rank_order(values, FREQS, 100.0, 15000.0)

    wb = load_workbook(xlsx)
    meta = {row[0].value: row[1].value for row in wb["curve_meta"].iter_rows(min_row=2)}
    assert meta["outlier_gate"] == "confirmed"
    assert meta["golden"] == expected_order[0]
    assert meta["aux"] == expected_order[1]  # aux_count=1 -> second-ranked sample

    headers, rows = _sheet_rows(xlsx, "curve_rank")
    by_sample = {r[headers.index("sample")]: r for r in rows}
    assert by_sample[expected_order[0]][headers.index("rank")] == 1
    assert by_sample[expected_order[0]][headers.index("role")] == "golden"
    assert by_sample[expected_order[1]][headers.index("role")] == "aux"
    # openpyxl round-trips an empty-string cell as None
    assert by_sample[expected_order[2]][headers.index("role")] in ("", None)


def test_excluded_sample_absent_from_mean_after_but_traceable_in_dev(tmp_path):
    xlsx, md = _make_batch(
        tmp_path,
        {
            "S1": [0.0, 0.0, 0.0, 0.0],
            "S2": [0.0, 0.0, 0.0, 0.0],
            "BAD": [50.0, 0.0, -50.0, 0.0],
        },
    )
    run(xlsx, aux_count=1, f_lo=100.0, f_hi=15000.0, exclude_arg="BAD", process_md=md)

    wb = load_workbook(xlsx)
    rank_headers, rank_rows = _sheet_rows(xlsx, "curve_rank")
    ranked_samples = {r[rank_headers.index("sample")] for r in rank_rows}
    assert "BAD" not in ranked_samples
    assert ranked_samples == {"S1", "S2"}

    dev_headers, dev_rows = _sheet_rows(xlsx, "curve_step3_dev")
    dev_sample_cols = dev_headers[1:]
    assert "BAD" in dev_sample_cols  # still traceable, per docstring intent

    dec_headers, dec_rows = _sheet_rows(xlsx, "outlier_decision")
    dec_by_sample = {r[dec_headers.index("sample")]: r for r in dec_rows}
    assert dec_by_sample["BAD"][dec_headers.index("in_mean_after")] == "no"
    assert dec_by_sample["S1"][dec_headers.index("in_mean_after")] == "yes"


def test_unknown_exclude_name_raises(tmp_path):
    xlsx, md = _make_batch(
        tmp_path, {"S1": [0.0, 0.0, 0.0, 0.0], "S2": [0.0, 0.0, 0.0, 0.0]}
    )
    with pytest.raises(SystemExit):
        run(xlsx, aux_count=1, f_lo=100.0, f_hi=15000.0, exclude_arg="GHOST", process_md=md)


def test_all_samples_excluded_raises(tmp_path):
    xlsx, md = _make_batch(
        tmp_path, {"S1": [0.0, 0.0, 0.0, 0.0], "S2": [0.0, 0.0, 0.0, 0.0]}
    )
    with pytest.raises(SystemExit):
        run(xlsx, aux_count=1, f_lo=100.0, f_hi=15000.0, exclude_arg="S1,S2", process_md=md)


def test_aux_count_controls_number_of_aux_roles(tmp_path):
    xlsx, md = _make_batch(
        tmp_path,
        {
            "S1": [0.0, 0.0, 0.0, 0.0],
            "S2": [0.5, 0.0, -0.5, 0.0],
            "S3": [1.0, 0.0, -1.0, 0.0],
            "S4": [1.5, 0.0, -1.5, 0.0],
        },
    )
    run(xlsx, aux_count=2, f_lo=100.0, f_hi=15000.0, exclude_arg="none", process_md=md)

    headers, rows = _sheet_rows(xlsx, "curve_rank")
    aux_roles = [r for r in rows if r[headers.index("role")] == "aux"]
    assert len(aux_roles) == 2


def test_high_q_exempt_sample_not_forced_out_of_ranking_by_script(tmp_path):
    """A high-Q spike sample gets suggested-exclude=False (exempt) in outlier_review,
    but the *script* never force-removes anyone — removal only happens via exclude_arg.
    This test locks in that gate-pending output still lists the exempt sample in
    outlier_review with exempt_high_q=yes and suggest_exclude=no.

    The spike is centered at 2000Hz (not 1000Hz): run_curves.py levels every curve by
    its own nearest-1000Hz value first, so a peak placed exactly at 1000Hz would get
    subtracted away by the leveling step before outlier detection ever sees it."""
    import math

    def log_freqs(n: int, f_lo: float, f_hi: float) -> list[float]:
        log_lo, log_hi = math.log10(f_lo), math.log10(f_hi)
        return [10 ** (log_lo + (log_hi - log_lo) * i / (n - 1)) for i in range(n)]

    freqs = log_freqs(200, 100.0, 15000.0)
    f0 = 2000.0
    half_bw = f0 / 40.0 / 2.0  # intrinsic Q=40 leaves margin above q_threshold=10
    # after averaging into mean_before with 3 quiet samples, the measured Q dilutes
    # to ~19 (verified numerically) — still comfortably >= q_threshold
    spike = [0.05 + 5.0 / (1.0 + ((f - f0) / half_bw) ** 2) for f in freqs]
    quiet = [0.05] * len(freqs)

    xlsx = tmp_path / "process.xlsx"
    md = tmp_path / "process.md"
    build_curves_workbook(
        xlsx,
        samples=["spike", "quiet1", "quiet2", "quiet3"],
        freqs=freqs,
        values_by_sample={
            "spike": spike,
            "quiet1": quiet,
            "quiet2": quiet,
            "quiet3": quiet,
        },
    )
    write_process_md(md)

    run(xlsx, aux_count=1, f_lo=100.0, f_hi=15000.0, exclude_arg=None, process_md=md)

    headers, rows = _sheet_rows(xlsx, "outlier_review")
    by_sample = {r[headers.index("sample")]: r for r in rows}
    assert by_sample["spike"][headers.index("exempt_high_q")] == "yes"
    assert by_sample["spike"][headers.index("suggest_exclude")] == "no"

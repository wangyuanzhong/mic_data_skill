# -*- coding: utf-8 -*-
"""
Read process.xlsx sheet `curves`, compute 1000 Hz sensitivity bins.
Inputs from params.json; writes analysis sheets + patches process.md log.
"""
from __future__ import annotations

import argparse
import math
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

sys.path.insert(0, str(Path(__file__).resolve().parent))
from params_io import load_params, patch_process_md_section, resolve_under_output  # noqa: E402


def _at_1000_nearest(
    freqs: list[float], vals: list[float]
) -> tuple[float, float, str]:
    """Return (freq_used, level, note)."""
    pairs = [(f, v) for f, v in zip(freqs, vals) if f is not None and v is not None]
    if not pairs:
        raise ValueError("empty column")
    target = 1000.0
    best_f, best_v = min(pairs, key=lambda p: abs(p[0] - target))
    note = "exact" if abs(best_f - target) < 1e-9 else "nearest"
    return float(best_f), float(best_v), note


def _bin_index(delta: float, width: float = 0.5) -> int:
    return math.floor(delta / width)


def _replace_sheet(wb, name: str) -> Worksheet:
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def _write_process_md(
    process_md: Path,
    *,
    midline_mode: str,
    mid: float,
    unit: str,
    golden_name: str,
    samples: list[str],
    levels: list[float],
    width: float,
) -> None:
    rows_md: list[str] = []
    for name, level in zip(samples, levels):
        delta = level - mid
        b = _bin_index(delta, width)
        role = "golden" if name == golden_name else "—"
        rows_md.append(f"| {name} | {level} | {delta} | {b} | {role} |")
    g_delta = ""
    g_bin = ""
    if golden_name:
        g_delta = levels[samples.index(golden_name)] - mid
        g_bin = _bin_index(float(g_delta), width)
    body = "\n".join(
        [
            "## 灵敏度分析",
            "",
            f"- 中线方式：{midline_mode}",
            f"- 中线数值：{mid} {unit}",
            f"- 灵敏度金标：{golden_name}（相对中线 {g_delta} dB，档位 {g_bin}）",
            f"- 样机数：{len(samples)}",
            "",
            f"| 样机 | 1000Hz ({unit}) | 相对中线 (dB) | 档位 | 角色 |",
            "|------|----------------|---------------|------|------|",
            *rows_md,
            "",
        ]
    )
    patch_process_md_section(process_md, "灵敏度分析", body)


def run(
    xlsx: Path,
    midline_mode: str,
    midline_db: float | None,
    unit: str,
    process_md: Path | None = None,
) -> None:
    wb = load_workbook(xlsx)
    if "curves" not in wb.sheetnames:
        raise SystemExit(f"missing sheet 'curves' in {xlsx}")
    ws = wb["curves"]

    samples: list[str] = []
    col = 2
    while True:
        name = ws.cell(1, col).value
        if name is None or str(name).strip() == "":
            break
        samples.append(str(name).strip())
        col += 1
    if not samples:
        raise SystemExit("no sample names on row 1 starting at column B")

    freqs: list[float] = []
    columns: list[list[float | None]] = [[] for _ in samples]
    row = 2
    while True:
        f = ws.cell(row, 1).value
        if f is None:
            break
        freqs.append(float(f))
        for i in range(len(samples)):
            raw = ws.cell(row, i + 2).value
            columns[i].append(float(raw) if raw is not None and raw != "" else None)
        row += 1
    if not freqs:
        raise SystemExit("no frequency rows in curves")

    freq_used: list[float] = []
    levels: list[float] = []
    notes: list[str] = []
    for col_vals in columns:
        f_use: list[float] = []
        v_use: list[float] = []
        for f, v in zip(freqs, col_vals):
            if v is not None:
                f_use.append(f)
                v_use.append(v)
        fu, level, note = _at_1000_nearest(f_use, v_use)
        freq_used.append(fu)
        levels.append(level)
        notes.append(note)

    pick = _replace_sheet(wb, "sens_step1_pick1000")
    pick.append(
        [
            "sample",
            "target_hz",
            "freq_used_hz",
            "abs_freq_error_hz",
            "level_at_freq_used",
            "pick_note",
            "unit",
        ]
    )
    for name, fu, level, note in zip(samples, freq_used, levels, notes):
        pick.append([name, 1000.0, fu, abs(fu - 1000.0), level, note, unit])

    if midline_mode == "mean":
        mid = sum(levels) / len(levels)
    elif midline_mode == "value":
        if midline_db is None:
            raise SystemExit("midline_db required when midline_mode=value")
        mid = float(midline_db)
    else:
        raise SystemExit("midline_mode must be mean or value")

    mid_ws = _replace_sheet(wb, "sens_step2_midline")
    mid_ws.append(["key", "value"])
    mid_ws.append(["midline_mode", midline_mode])
    mid_ws.append(["midline_db", mid])
    mid_ws.append(["unit", unit])
    mid_ws.append(["n_samples", len(samples)])
    mid_ws.append(["bin_width_db", 0.5])
    mid_ws.append([])
    mid_ws.append(
        [
            "sample",
            "level_1000",
            "included_in_mean",
            "contribution_note",
        ]
    )
    for name, level in zip(samples, levels):
        if midline_mode == "mean":
            mid_ws.append([name, level, "yes", "average of all level_1000"])
        else:
            mid_ws.append([name, level, "no", "midline from params midline_db"])

    width = 0.5
    ranked = sorted(
        zip(samples, levels, notes, freq_used),
        key=lambda t: (abs(t[1] - mid), t[0]),
    )
    golden_name = ranked[0][0] if ranked else ""
    order = {name: i for i, (name, *_rest) in enumerate(ranked, start=1)}

    detail = _replace_sheet(wb, "sens_step3_bins")
    detail.append(
        [
            "sample",
            "freq_used_hz",
            "level_1000",
            "midline_db",
            "delta_to_mid",
            "abs_delta_to_mid",
            "bin",
            "bin_lo_db",
            "bin_hi_db",
            "pick_note",
            "abs_delta_rank",
            "role",
            "unit",
        ]
    )
    for name, level, note, fu in zip(samples, levels, notes, freq_used):
        delta = level - mid
        b = _bin_index(delta, width)
        role = "golden" if name == golden_name else ""
        detail.append(
            [
                name,
                fu,
                level,
                mid,
                delta,
                abs(delta),
                b,
                b * width,
                (b + 1) * width,
                note,
                order[name],
                role,
                unit,
            ]
        )

    sens = _replace_sheet(wb, "sensitivity")
    sens.append(
        [
            "sample",
            "freq_used_hz",
            "level_1000",
            "delta_to_mid",
            "bin",
            "bin_lo_db",
            "bin_hi_db",
            "note",
            "role",
            "unit",
        ]
    )
    for name, level, note, fu in zip(samples, levels, notes, freq_used):
        delta = level - mid
        b = _bin_index(delta, width)
        role = "golden" if name == golden_name else ""
        sens.append(
            [
                name,
                fu,
                level,
                delta,
                b,
                b * width,
                (b + 1) * width,
                note,
                role,
                unit,
            ]
        )

    meta = _replace_sheet(wb, "sensitivity_meta")
    meta.append(["key", "value"])
    meta.append(["midline_mode", midline_mode])
    meta.append(["midline_db", mid])
    meta.append(["unit", unit])
    meta.append(["n_samples", len(samples)])
    meta.append(["bin_width_db", width])
    meta.append(["golden", golden_name])
    if golden_name:
        g_level = levels[samples.index(golden_name)]
        g_delta = g_level - mid
        meta.append(["golden_delta_to_mid", g_delta])
        meta.append(["golden_bin", _bin_index(g_delta, width)])
    meta.append(
        [
            "process_sheets",
            "sens_step1_pick1000 | sens_step2_midline | sens_step3_bins | sensitivity",
        ]
    )

    wb.save(xlsx)

    if process_md is not None:
        _write_process_md(
            process_md,
            midline_mode=midline_mode,
            mid=mid,
            unit=unit,
            golden_name=golden_name,
            samples=samples,
            levels=levels,
            width=width,
        )

    print(f"OK wrote sensitivity process sheets -> {xlsx}")
    if process_md is not None:
        print(f"OK patched process.md section 灵敏度分析 -> {process_md}")
    print(
        f"midline_mode={midline_mode} midline_db={mid:.6g} unit={unit} "
        f"n={len(samples)} golden={golden_name}"
    )


def main() -> None:
    p = argparse.ArgumentParser(
        description="1000Hz sensitivity bins; inputs from params.json"
    )
    p.add_argument(
        "--params",
        type=Path,
        required=True,
        help="Path to params.json (deterministic inputs)",
    )
    args = p.parse_args()
    params = load_params(args.params)
    xlsx = resolve_under_output(params, "process_xlsx", args.params)
    md = resolve_under_output(params, "process_md", args.params)
    if not xlsx.is_file():
        raise SystemExit(f"file not found: {xlsx}")
    if not md.is_file():
        raise SystemExit(f"process.md not found: {md}")

    sens = params.get("sensitivity") or {}
    midline_mode = sens.get("midline_mode") or "mean"
    midline_db = sens.get("midline_db")
    unit = params.get("unit") or "dB"
    if midline_mode not in ("mean", "value"):
        raise SystemExit("sensitivity.midline_mode must be mean or value")
    if midline_mode == "value" and midline_db is None:
        raise SystemExit("sensitivity.midline_db required when midline_mode=value")

    run(
        xlsx,
        midline_mode,
        float(midline_db) if midline_db is not None else None,
        unit,
        process_md=md,
    )


if __name__ == "__main__":
    main()

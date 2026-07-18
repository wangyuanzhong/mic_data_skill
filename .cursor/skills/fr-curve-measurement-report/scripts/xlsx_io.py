# -*- coding: utf-8 -*-
"""Read process.xlsx sheets for report generation (numerics quantized to 1 dp)."""
from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from quantize import q1


def read_meta_kv(ws: Worksheet) -> dict[str, str]:
    """Read key/value pairs from columns A/B (skip non-pairs)."""
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        if key.lower() in ("key", ""):
            continue
        val = "" if row[1] is None else str(row[1]).strip()
        out[key] = val
    return out


def read_wide_curves(
    ws: Worksheet,
) -> tuple[list[float], list[str], list[list[float | None]]]:
    """Wide layout: A=freq, row1=sample names, body=amplitudes. Values → 1 dp."""
    samples: list[str] = []
    for col in range(2, ws.max_column + 1):
        name = ws.cell(1, col).value
        if name is None or str(name).strip() == "":
            break
        samples.append(str(name).strip())
    if not samples:
        raise ValueError("no sample columns in wide sheet")

    freqs: list[float] = []
    columns: list[list[float | None]] = [[] for _ in samples]
    for row in range(2, ws.max_row + 1):
        f = ws.cell(row, 1).value
        if f is None or f == "":
            continue
        fq = q1(f)
        assert fq is not None
        freqs.append(fq)
        for i in range(len(samples)):
            raw = ws.cell(row, i + 2).value
            columns[i].append(
                q1(raw) if raw is not None and raw != "" else None
            )
    return freqs, samples, columns


def read_table_dicts(ws: Worksheet) -> list[dict[str, object]]:
    """First row headers → list of row dicts; numeric cells → 1 dp."""
    headers: list[str] = []
    for col in range(1, ws.max_column + 1):
        h = ws.cell(1, col).value
        if h is None:
            break
        headers.append(str(h).strip())
    rows: list[dict[str, object]] = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is None and all(
            ws.cell(r, c).value is None for c in range(1, len(headers) + 1)
        ):
            continue
        item: dict[str, object] = {}
        empty = True
        for c, h in enumerate(headers, start=1):
            v = ws.cell(r, c).value
            if v is not None and v != "":
                empty = False
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                item[h] = q1(v)
            else:
                item[h] = v
        if not empty:
            rows.append(item)
    return rows

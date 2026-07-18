# -*- coding: utf-8 -*-
"""Report readiness gates on analysis output directory."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from xlsx_io import read_meta_kv

REQUIRED_SHEETS = (
    "curves",
    "sensitivity",
    "sensitivity_meta",
    "curve_step1_leveled",
    "curve_mean_before_outlier",
    "curve_mean_after_outlier",
    "curve_step3_dev",
    "outlier_decision",
    "curve_rank",
    "curve_step4_envelope",
    "curve_meta",
)


def check_report_ready(output_dir: Path) -> None:
    output_dir = Path(output_dir)
    params_path = output_dir / "params.json"
    xlsx_path = output_dir / "process.xlsx"
    if not params_path.is_file():
        raise SystemExit(f"缺少 params.json：{params_path}")
    if not xlsx_path.is_file():
        raise SystemExit(f"缺少 process.xlsx：{xlsx_path}")

    try:
        params = json.loads(params_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        raise SystemExit(f"params.json 无法解析：{e}") from e
    if not isinstance(params, dict):
        raise SystemExit("params.json 必须是对象")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if "curve_meta" not in wb.sheetnames:
            raise SystemExit("process.xlsx 缺少 sheet：curve_meta")
        meta = read_meta_kv(wb["curve_meta"])
        gate = meta.get("outlier_gate", "").strip().lower()
        if gate != "confirmed":
            raise SystemExit(
                f"outlier_gate 未确认（当前={meta.get('outlier_gate', '(空)')}），"
                "请先完成选标奇异值确认后再出报告"
            )
        missing = [n for n in REQUIRED_SHEETS if n not in wb.sheetnames]
        if missing:
            raise SystemExit(f"process.xlsx 缺少 sheet：{', '.join(missing)}")
    finally:
        wb.close()

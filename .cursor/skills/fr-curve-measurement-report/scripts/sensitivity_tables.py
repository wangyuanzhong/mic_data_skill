# -*- coding: utf-8 -*-
"""Build sensitivity HTML fragments from process.xlsx (deterministic)."""
from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from quantize import q1
from xlsx_io import read_meta_kv, read_table_dicts


def _fmt_db(v: object) -> str:
    try:
        x = q1(v)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return str(v)
    if x is None:
        return ""
    if abs(x - round(x)) < 1e-9:
        return str(int(round(x)))
    return f"{x:.1f}"


def _midline_from_rows(rows: list[dict], meta: dict[str, str]) -> float | None:
    raw = meta.get("midline_db")
    if raw not in (None, ""):
        try:
            return float(raw)
        except ValueError:
            pass
    for r in rows:
        level = r.get("level_1000")
        delta = r.get("delta_to_mid")
        if level is not None and delta is not None:
            return float(level) - float(delta)
    return None


def count_sensitivity_bins_absolute(
    rows: list[dict], midline: float
) -> list[tuple[float, float, int]]:
    """
    Count samples per sensitivity bin, expressed as absolute amplitude limits:
    abs_lo = midline + bin_lo_db, abs_hi = midline + bin_hi_db.
    """
    counter: Counter[tuple[float, float]] = Counter()
    for r in rows:
        lo = r.get("bin_lo_db")
        hi = r.get("bin_hi_db")
        if lo is None or hi is None:
            continue
        abs_lo = midline + float(lo)
        abs_hi = midline + float(hi)
        counter[(q1(abs_lo) or abs_lo, q1(abs_hi) or abs_hi)] += 1
    return [
        (lo, hi, n)
        for (lo, hi), n in sorted(counter.items(), key=lambda t: t[0][0])
    ]


def _load_sensitivity(output_dir: Path) -> tuple[list[dict], dict[str, str]]:
    xlsx = Path(output_dir) / "process.xlsx"
    if not xlsx.is_file():
        return [], {}
    wb = load_workbook(xlsx, data_only=True)
    try:
        if "sensitivity" not in wb.sheetnames:
            return [], {}
        rows = read_table_dicts(wb["sensitivity"])
        meta: dict[str, str] = {}
        if "sensitivity_meta" in wb.sheetnames:
            meta = read_meta_kv(wb["sensitivity_meta"])
        return rows, meta
    finally:
        wb.close()


def sensitivity_detail_html(output_dir: Path) -> str:
    """Per-sample sensitivity table from process.xlsx (Agent must not hand-write this)."""
    rows, meta = _load_sensitivity(output_dir)
    if not rows:
        return ""
    unit = ""
    if rows[0].get("unit"):
        unit = str(rows[0]["unit"])
    elif meta.get("unit"):
        unit = str(meta["unit"])
    level_h = f"1000Hz ({unit})" if unit else "1000Hz"
    lines = [
        "<h3>灵敏度明细</h3>",
        "<table>",
        "<thead><tr>"
        "<th>样机</th>"
        f"<th>{level_h}</th>"
        "<th>相对中线 (dB)</th>"
        "<th>档位</th>"
        "<th>角色</th>"
        "</tr></thead>",
        "<tbody>",
    ]
    for r in rows:
        role = r.get("role") or "—"
        if role == "golden":
            role = "金标"
        lines.append(
            "<tr>"
            f"<td>{r.get('sample', '')}</td>"
            f"<td>{_fmt_db(r.get('level_1000'))}</td>"
            f"<td>{_fmt_db(r.get('delta_to_mid'))}</td>"
            f"<td>{r.get('bin', '')}</td>"
            f"<td>{role}</td>"
            "</tr>"
        )
    lines.extend(["</tbody>", "</table>"])
    return "\n".join(lines)


def sensitivity_bin_count_html(output_dir: Path) -> str:
    """HTML table: 幅度范围 (A~B) / 样机数 — absolute dB, not relative to midline."""
    rows, meta = _load_sensitivity(output_dir)
    if not rows:
        return ""
    midline = _midline_from_rows(rows, meta)
    if midline is None:
        return ""
    bins = count_sensitivity_bins_absolute(rows, midline)
    if not bins:
        return ""
    unit = ""
    if rows[0].get("unit"):
        unit = f" ({rows[0]['unit']})"
    elif meta.get("unit"):
        unit = f" ({meta['unit']})"
    lines = [
        "<h3>灵敏度分档计数</h3>",
        "<table>",
        f"<thead><tr><th>幅度范围{unit}</th><th>样机数</th></tr></thead>",
        "<tbody>",
    ]
    for lo, hi, n in bins:
        lines.append(
            f"<tr><td>{_fmt_db(lo)}~{_fmt_db(hi)}</td><td>{n}</td></tr>"
        )
    lines.extend(["</tbody>", "</table>"])
    return "\n".join(lines)

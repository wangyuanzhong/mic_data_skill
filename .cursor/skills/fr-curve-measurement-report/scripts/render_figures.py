# -*- coding: utf-8 -*-
"""Render report figures from params.json + process.xlsx."""
from __future__ import annotations

import argparse
import json
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import plot_style  # noqa: F401 — sets CJK font before pyplot use
import matplotlib.pyplot as plt  # noqa: E402
from openpyxl import load_workbook

from consistency import compute_sigma_curve, summarize_consistency_zh
from envelope_bins import bin_samples_by_envelope
from gates import check_report_ready
from plot_style import (
    REPORT_F_MAX_HZ,
    REPORT_F_MIN_HZ,
    clamp_sigma_ylim,
    draw_analysis_band,
    draw_envelope_box,
    prepare_curve_for_display,
    setup_log_freq_axis,
)
from utf8_boot import ensure_utf8_stdio, read_text, write_text
from xlsx_io import read_meta_kv, read_table_dicts, read_wide_curves


def _load_params(output_dir: Path) -> dict:
    return json.loads(read_text(output_dir / "params.json"))


def _series_map(
    samples: list[str], columns: list[list[float | None]]
) -> dict[str, list[float | None]]:
    return {s: columns[i] for i, s in enumerate(samples)}


def _apply_freq_axis(ax, freqs: list[float], f_lo: float, f_hi: float) -> None:
    # 报告横轴固定 20–20kHz；分析上下限只画虚线+标注
    del freqs  # data span does not drive axis limits
    setup_log_freq_axis(ax, REPORT_F_MIN_HZ, REPORT_F_MAX_HZ)
    draw_analysis_band(ax, f_lo, f_hi)


def _save_fig(fig, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.tight_layout()
    fig.subplots_adjust(bottom=0.20)
    fig.savefig(path, dpi=150)
    plt.close(fig)
    return path


def _plot_overlay(
    path: Path,
    freqs: list[float],
    series: dict[str, list[float | None]],
    title: str,
    ylabel: str,
    f_lo: float,
    f_hi: float,
    mean_curve: list[float | None] | None = None,
    mean_label: str = "均值",
) -> Path:
    fig, ax = plt.subplots(figsize=(10, 5.5))
    for name, amps in series.items():
        xf, ys = prepare_curve_for_display(freqs, amps)
        ax.plot(xf, ys, linewidth=1.1, label=name)
    if mean_curve is not None:
        xf, ys = prepare_curve_for_display(freqs, mean_curve)
        ax.plot(xf, ys, color="black", linewidth=2.0, label=mean_label)
    _apply_freq_axis(ax, freqs, f_lo, f_hi)
    ax.set_ylabel(ylabel)
    ax.set_title(title)
    if len(series) <= 12:
        ax.legend(fontsize=7, loc="best")
    return _save_fig(fig, path)


def _read_mean_curve(ws) -> tuple[list[float], list[float | None]]:
    freqs: list[float] = []
    mean: list[float | None] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        freqs.append(float(row[0]))
        mean.append(None if row[1] is None else float(row[1]))
    return freqs, mean


def render_all_figures(
    output_dir: Path,
    figures_dir: Path | None = None,
) -> list[Path]:
    output_dir = Path(output_dir)
    check_report_ready(output_dir)
    params = _load_params(output_dir)
    f_lo = float(params["f_lo_hz"])
    f_hi = float(params["f_hi_hz"])
    unit = str(params.get("unit") or "dB")
    figures_dir = Path(figures_dir) if figures_dir else output_dir / "figures"
    figures_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(output_dir / "process.xlsx", data_only=True)
    freqs, samples, leveled_cols = read_wide_curves(wb["curve_step1_leveled"])
    leveled = _series_map(samples, leveled_cols)
    _, _, dev_cols = read_wide_curves(wb["curve_step3_dev"])
    devs = _series_map(samples, dev_cols)
    _mf, mean_before = _read_mean_curve(wb["curve_mean_before_outlier"])
    _mf2, mean_after = _read_mean_curve(wb["curve_mean_after_outlier"])

    decisions = read_table_dicts(wb["outlier_decision"])
    excluded = [
        str(r["sample"])
        for r in decisions
        if str(r.get("user_exclude", "")).lower() in ("yes", "true", "1")
    ]
    kept = [
        str(r["sample"])
        for r in decisions
        if str(r.get("in_mean_after", "")).lower() in ("yes", "true", "1")
    ]
    if not kept:
        kept = [s for s in samples if s not in excluded]

    rank_rows = read_table_dicts(wb["curve_rank"])
    # normalize delta key
    for r in rank_rows:
        if "delta" not in r and "delta_db" in r:
            r["delta"] = r["delta_db"]
    meta = read_meta_kv(wb["curve_meta"])
    golden = meta.get("golden") or ""
    for r in rank_rows:
        if str(r.get("role", "")).lower() == "golden":
            golden = str(r["sample"])
            break
    aux = [
        str(r["sample"])
        for r in rank_rows
        if str(r.get("role", "")).lower() == "aux"
    ]

    out_paths: list[Path] = []

    # 3: outliers vs mean_before
    outlier_series = {n: leveled[n] for n in excluded if n in leveled}
    p = figures_dir / "奇异值与剔异前均值.png"
    title = "奇异值与剔异前均值（1 kHz 归零）"
    if not outlier_series:
        title += "（本批无剔除）"
    _plot_overlay(
        p,
        freqs,
        outlier_series,
        title,
        f"Level ({unit}, leveled)",
        f_lo,
        f_hi,
        mean_curve=mean_before,
        mean_label="剔异前均值",
    )
    out_paths.append(p)

    # 2: kept leveled overlay（标题不再强调「剔异后」）
    p = figures_dir / "归零叠图.png"
    _plot_overlay(
        p,
        freqs,
        {n: leveled[n] for n in kept if n in leveled},
        "归零叠图",
        f"Level ({unit}, leveled)",
        f_lo,
        f_hi,
        mean_curve=mean_after,
        mean_label="均值",
    )
    out_paths.append(p)

    # 1: envelope bins
    bins = bin_samples_by_envelope(rank_rows)
    for shell, names in bins.items():
        # filename: 包络分档_正负0.5dB.png
        shell_label = f"{shell:g}"
        p = figures_dir / f"包络分档_正负{shell_label}dB.png"
        fig, ax = plt.subplots(figsize=(10, 5.5))
        for name in names:
            if name not in devs:
                continue
            xf, ys = prepare_curve_for_display(freqs, devs[name])
            ax.plot(xf, ys, linewidth=1.1, label=name)
        _apply_freq_axis(ax, freqs, f_lo, f_hi)
        draw_envelope_box(ax, f_lo, f_hi, float(shell))
        ax.set_ylabel("Deviation vs mean (dB)")
        ax.set_title(f"包络分档 ±{shell_label} dB（归零后相对均值）")
        ax.legend(fontsize=7, loc="best")
        _save_fig(fig, p)
        out_paths.append(p)

    # 4: golden abs + golden-mean
    if golden and golden in leveled:
        p = figures_dir / "金标绝对与偏差.png"
        fig, ax = plt.subplots(figsize=(10, 5.5))
        g_abs_raw = leveled[golden]
        g_dev_raw: list[float | None] = []
        for a, m in zip(leveled[golden], mean_after):
            if a is None or m is None:
                g_dev_raw.append(None)
            else:
                g_dev_raw.append(float(a) - float(m))
        xf_a, g_abs = prepare_curve_for_display(freqs, g_abs_raw)
        xf_d, g_dev = prepare_curve_for_display(freqs, g_dev_raw)
        ax.plot(xf_a, g_abs, linewidth=1.6, label="金标绝对")
        ax.plot(xf_d, g_dev, linewidth=1.6, label="金标−均值")
        _apply_freq_axis(ax, freqs, f_lo, f_hi)
        ax.set_ylabel(f"dB ({unit} / deviation)")
        ax.set_title(f"金标绝对与偏差（{golden}）")
        ax.legend(loc="best")
        _save_fig(fig, p)
        out_paths.append(p)

    # 5: aux deviation
    if aux:
        p = figures_dir / "辅标偏差叠图.png"
        _plot_overlay(
            p,
            freqs,
            {n: devs[n] for n in aux if n in devs},
            "辅标偏差叠图（相对均值）",
            "Deviation (dB)",
            f_lo,
            f_hi,
        )
        out_paths.append(p)

        # 6: aux absolute
        p = figures_dir / "辅标绝对叠图.png"
        _plot_overlay(
            p,
            freqs,
            {n: leveled[n] for n in aux if n in leveled},
            "辅标绝对叠图（1 kHz 归零）",
            f"Level ({unit}, leveled)",
            f_lo,
            f_hi,
        )
        out_paths.append(p)

    # 7: consistency sigma
    amps_f: dict[str, list[float]] = {}
    for n in kept:
        if n not in leveled:
            continue
        amps_f[n] = [
            0.0 if v is None else float(v) for v in leveled[n]
        ]
    if len(amps_f) >= 2:
        sigma = compute_sigma_curve(freqs, amps_f, f_lo, f_hi)
        note = summarize_consistency_zh(sigma)
        note_path = figures_dir / "批量一致性说明.txt"
        write_text(note_path, note + "\n")

        p = figures_dir / "批量一致性sigma.png"
        fig, ax = plt.subplots(figsize=(10, 5.2))
        xs, ys = [], []
        for f, s in zip(sigma["freqs_hz"], sigma["sigma_db"]):
            if s is None:
                if xs:
                    xf, ys_s = prepare_curve_for_display(xs, ys)
                    # Line only (no under-curve fill / glow); match directivity σ style
                    ax.plot(xf, ys_s, linewidth=1.2)
                    xs, ys = [], []
                continue
            xs.append(f)
            ys.append(s)
        if xs:
            xf, ys_s = prepare_curve_for_display(xs, ys)
            ax.plot(xf, ys_s, linewidth=1.2)
        _apply_freq_axis(ax, freqs, f_lo, f_hi)
        ax.set_ylabel("σ (dB)")
        ax.set_title("批量一致性 σ(f)")
        clamp_sigma_ylim(ax)
        _save_fig(fig, p)
        out_paths.append(p)
        out_paths.append(note_path)

    wb.close()
    return out_paths


def main() -> None:
    ensure_utf8_stdio()
    p = argparse.ArgumentParser(description="Render FR measurement report figures")
    p.add_argument("--params", required=True, help="Path to params.json")
    args = p.parse_args()
    params_path = Path(args.params)
    params = json.loads(read_text(params_path))
    output_dir = Path(params.get("output_dir") or params_path.parent)
    paths = render_all_figures(output_dir)
    print(f"OK wrote {len(paths)} figures -> {output_dir / 'figures'}")
    for path in paths:
        print(path.name)


if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""Synthetic analysis output fixtures for report skill tests."""
from __future__ import annotations

import json
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

SCRIPTS_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(SCRIPTS_DIR))


FREQS = [100.0, 500.0, 1000.0, 5000.0, 10000.0]
SAMPLES = ["S1", "S2", "S3", "S4"]


def _write_wide(ws, freqs, samples, columns, freq_header="freq_hz"):
    ws.append([freq_header, *samples])
    for r, f in enumerate(freqs):
        ws.append([f, *[columns[i][r] for i in range(len(samples))]])


def _write_mean(ws, freqs, mean):
    ws.append(["freq_hz", "mean_leveled_db", "n_samples_at_freq"])
    for f, m in zip(freqs, mean):
        ws.append([f, m, len(SAMPLES)])


def _leveled_columns():
    # Slightly different shapes around mean 0 after level
    return [
        [0.0, 0.1, 0.0, -0.2, 0.1],
        [0.1, 0.0, 0.0, 0.1, -0.1],
        [0.5, 0.4, 0.0, 0.6, 0.5],  # larger → outlier candidate
        [-0.1, 0.0, 0.0, -0.1, 0.0],
    ]


def build_ready_workbook(path: Path, *, pending: bool = False) -> None:
    leveled = _leveled_columns()
    mean_before = [
        sum(leveled[i][r] for i in range(len(SAMPLES))) / len(SAMPLES)
        for r in range(len(FREQS))
    ]
    # S3 excluded when confirmed
    kept_idx = [0, 1, 3] if not pending else list(range(4))
    mean_after = [
        sum(leveled[i][r] for i in kept_idx) / len(kept_idx)
        for r in range(len(FREQS))
    ]
    dev = [
        [leveled[i][r] - mean_after[r] for r in range(len(FREQS))]
        for i in range(len(SAMPLES))
    ]

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "curves"
    _write_wide(ws0, FREQS, SAMPLES, leveled)

    sens = wb.create_sheet("sensitivity")
    sens.append(
        [
            "sample",
            "freq_used_hz",
            "level_1000",
            "delta_to_mid",
            "bin",
            "bin_lo_db",
            "bin_hi_db",
            "note",
            "role",
            "unit",
        ]
    )
    for i, name in enumerate(SAMPLES):
        role = "golden" if name == "S1" else ""
        sens.append([name, 1000.0, -40.0 + i * 0.1, i * 0.1, 0, 0.0, 0.5, "", role, "dBV"])

    sm = wb.create_sheet("sensitivity_meta")
    sm.append(["key", "value"])
    sm.append(["midline_mode", "mean"])
    sm.append(["midline_db", -39.85])
    sm.append(["unit", "dBV"])
    sm.append(["n_samples", 4])
    sm.append(["bin_width_db", 0.5])
    sm.append(["golden", "S1"])

    lev = wb.create_sheet("curve_step1_leveled")
    _write_wide(lev, FREQS, SAMPLES, leveled)

    mb = wb.create_sheet("curve_mean_before_outlier")
    _write_mean(mb, FREQS, mean_before)

    ma = wb.create_sheet("curve_mean_after_outlier")
    _write_mean(ma, FREQS, mean_after)

    dev_ws = wb.create_sheet("curve_step3_dev")
    _write_wide(dev_ws, FREQS, SAMPLES, dev)

    if not pending:
        dec = wb.create_sheet("outlier_decision")
        dec.append(
            [
                "sample",
                "suggest_exclude",
                "exempt_high_q",
                "user_exclude",
                "in_mean_after",
                "decision_note",
            ]
        )
        for name in SAMPLES:
            ex = name == "S3"
            dec.append(
                [
                    name,
                    "yes" if ex else "no",
                    "no",
                    "yes" if ex else "no",
                    "no" if ex else "yes",
                    "已剔除" if ex else "保留",
                ]
            )

        rank = wb.create_sheet("curve_rank")
        rank.append(
            [
                "sample",
                "max_abs_dev",
                "freq_at_max_abs",
                "delta",
                "rms",
                "n_points_in_band",
                "rank",
                "role",
            ]
        )
        # kept: S1,S2,S4 — assign deltas
        rank_rows = [
            ("S1", 0.15, 5000.0, 0.2, 0.1, 5, 1, "golden"),
            ("S4", 0.2, 5000.0, 0.2, 0.12, 5, 2, "aux"),
            ("S2", 0.35, 100.0, 0.4, 0.2, 5, 3, "aux"),
        ]
        for row in rank_rows:
            rank.append(list(row))

        env = wb.create_sheet("curve_step4_envelope")
        env.append(
            [
                "sample",
                "max_abs_dev_db",
                "freq_at_max_hz",
                "delta_db",
                "rms_db",
                "n_points_in_band",
                "f_lo_hz",
                "f_hi_hz",
                "delta_step_db",
                "rank",
                "role",
                "excluded_from_mean",
            ]
        )
        for name, max_abs, freq_at, delta, rms, n, rnk, role in rank_rows:
            env.append(
                [name, max_abs, freq_at, delta, rms, n, 100, 15000, 0.1, rnk, role, "no"]
            )

    meta = wb.create_sheet("curve_meta")
    meta.append(["key", "value"])
    meta.append(["aux_count", 2])
    meta.append(["f_lo_hz", 100])
    meta.append(["f_hi_hz", 15000])
    meta.append(["n_samples_all", 4])
    if pending:
        meta.append(["outlier_gate", "pending"])
        meta.append(["excluded", "(pending user confirm)"])
    else:
        meta.append(["outlier_gate", "confirmed"])
        meta.append(["excluded", "S3"])
        meta.append(["n_samples_after_exclude", 3])
        meta.append(["golden", "S1"])
        meta.append(["aux", "S4, S2"])

    wb.save(path)


def _write_params(dir_path: Path, *, exclude) -> None:
    params = {
        "product": "TEST-MIC",
        "note": "合成夹具",
        "data_root": str(dir_path),
        "output_dir": str(dir_path),
        "unit": "dBV",
        "f_lo_hz": 100,
        "f_hi_hz": 15000,
        "process_xlsx": "process.xlsx",
        "process_md": "process.md",
        "sensitivity": {"midline_mode": "mean", "midline_db": None},
        "curves": {"aux_count": 2, "exclude": exclude, "q_threshold": 10},
    }
    (dir_path / "params.json").write_text(
        json.dumps(params, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


@pytest.fixture
def synthetic_ready_dir(tmp_path: Path) -> Path:
    build_ready_workbook(tmp_path / "process.xlsx", pending=False)
    _write_params(tmp_path, exclude=["S3"])
    return tmp_path


@pytest.fixture
def synthetic_pending_dir(tmp_path: Path) -> Path:
    build_ready_workbook(tmp_path / "process.xlsx", pending=True)
    _write_params(tmp_path, exclude=None)
    return tmp_path

# -*- coding: utf-8 -*-
"""
run_deltas.py — write delta_<tag> sheet per non-axial angle.

For each non-axial angle sheet, subtract the axial sheet column-by-column
(same sample order, same frequency grid) and write the result to a new
sheet named `delta_<normalize_angle_tag(angle)>`.

Usage:
    python run_deltas.py --params <output_dir>/params.json

Angle wide sheets in process.xlsx are looked up by
`angle_sheet_name(params label)` (= normalize_angle_tag), not the raw
params string. Example: params.angles=["0°","90°"] → sheets "0", "90".
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from angle_sheets import (
    angle_sheet_name,
    assert_columns_aligned,
    normalize_angle_tag,
    read_angle_sheet,
    subtract_columns,
)
from params_io import load_params, resolve_under_output
from utf8_boot import ensure_utf8_stdio


def _replace_sheet(wb, name: str) -> Worksheet:
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def _write_wide(
    ws: Worksheet,
    freqs: list[float],
    samples: list[str],
    columns: list[list[float | None]],
) -> None:
    ws.append(["freq_hz", *samples])
    for r, f in enumerate(freqs):
        ws.append([f, *[columns[i][r] for i in range(len(samples))]])


def main(argv: list[str] | None = None) -> int:
    ensure_utf8_stdio()
    parser = argparse.ArgumentParser(description="Write delta_<tag> sheets.")
    parser.add_argument("--params", required=True, type=Path)
    args = parser.parse_args(argv)

    params = load_params(args.params)
    xlsx_path = resolve_under_output(params, "process_xlsx", args.params)
    angles = params.get("angles") or []
    axial = params.get("axial_angle") or ""

    if not angles:
        print("params.angles is empty", file=sys.stderr)
        return 2
    if not axial:
        print("params.axial_angle is empty", file=sys.stderr)
        return 2
    if axial not in angles:
        print(f"axial_angle {axial!r} not in angles {angles!r}", file=sys.stderr)
        return 2

    wb = load_workbook(xlsx_path)
    axial_sheet = angle_sheet_name(axial)
    if axial_sheet not in wb.sheetnames:
        print(
            f"axial sheet {axial_sheet!r} (from {axial!r}) missing in {xlsx_path}",
            file=sys.stderr,
        )
        return 2

    axial_freqs, axial_samples, axial_cols = read_angle_sheet(wb[axial_sheet])

    written: list[str] = []
    for angle in angles:
        if angle == axial:
            continue
        sheet = angle_sheet_name(angle)
        if sheet not in wb.sheetnames:
            print(
                f"angle sheet {sheet!r} (from {angle!r}) missing in {xlsx_path}",
                file=sys.stderr,
            )
            return 2
        a_freqs, a_samples, a_cols = read_angle_sheet(wb[sheet])
        if a_samples != axial_samples:
            print(
                f"sample order mismatch in {angle!r}: {a_samples} vs axial {axial_samples}",
                file=sys.stderr,
            )
            return 2
        try:
            assert_columns_aligned(
                freqs_a=axial_freqs,
                cols_a=axial_cols,
                freqs_b=a_freqs,
                cols_b=a_cols,
            )
        except ValueError as e:
            print(f"alignment error for {angle!r}: {e}", file=sys.stderr)
            return 2
        delta_cols = subtract_columns(a_cols, axial_cols)
        tag = normalize_angle_tag(angle)
        sheet_name = f"delta_{tag}"
        ws = _replace_sheet(wb, sheet_name)
        _write_wide(ws, axial_freqs, axial_samples, delta_cols)
        written.append(sheet_name)

    wb.save(xlsx_path)
    print(f"run_deltas: wrote {len(written)} sheets: {', '.join(written)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

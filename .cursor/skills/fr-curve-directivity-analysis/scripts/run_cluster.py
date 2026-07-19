# -*- coding: utf-8 -*-
"""
run_cluster.py — Euclidean distance matrix + hierarchical clustering.

For each non-axial delta_<tag> sheet, band-limit to [f_lo_hz, f_hi_hz],
compute sample×sample Euclidean distances (None pairs skipped), choose k via
average-linkage silhouette, and write cluster_dist / cluster_suggest /
cluster_meta / cluster_final (if missing).

Usage:
    python run_cluster.py --params <output_dir>/params.json
"""
from __future__ import annotations

import argparse
import math
import sys
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from scipy.cluster.hierarchy import fcluster, linkage
from scipy.spatial.distance import squareform

from angle_sheets import normalize_angle_tag, read_angle_sheet
from params_io import load_params, resolve_under_output
from utf8_boot import ensure_utf8_stdio


def _replace_sheet(wb, name: str) -> Worksheet:
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def _band_mask(freqs: list[float], f_lo: float, f_hi: float) -> list[bool]:
    return [f_lo <= float(f) <= f_hi for f in freqs]


def _euclid(a: list[float | None], b: list[float | None]) -> float | None:
    """Euclidean distance; only positions where both sides are non-None.

    Returns None if no overlapping valid points.
    """
    if len(a) != len(b):
        raise ValueError("vector length mismatch in _euclid")
    s = 0.0
    n = 0
    for va, vb in zip(a, b):
        if va is None or vb is None:
            continue
        d = float(va) - float(vb)
        s += d * d
        n += 1
    if n == 0:
        return None
    return math.sqrt(s)


def _band_columns(
    freqs: list[float],
    cols: list[list[float | None]],
    f_lo: float,
    f_hi: float,
) -> list[list[float | None]]:
    mask = _band_mask(freqs, f_lo, f_hi)
    return [[col[i] for i in range(len(freqs)) if mask[i]] for col in cols]


def _is_clusterable(vec: list[float | None]) -> bool:
    return any(v is not None for v in vec)


def _class_center(members: list[list[float | None]]) -> list[float | None]:
    if not members:
        return []
    n_freq = len(members[0])
    out: list[float | None] = []
    for i in range(n_freq):
        vals = [m[i] for m in members if m[i] is not None]
        if not vals:
            out.append(None)
        else:
            out.append(sum(vals) / len(vals))
    return out


def _silhouette_precomputed(dist: np.ndarray, labels: np.ndarray) -> float:
    """Mean silhouette score using a precomputed pairwise distance matrix."""
    n = dist.shape[0]
    if n < 2:
        return 0.0
    labels = np.asarray(labels)
    unique = np.unique(labels)
    if len(unique) < 2:
        return 0.0

    scores: list[float] = []
    for i in range(n):
        same = labels == labels[i]
        n_same = int(np.sum(same)) - 1  # exclude self
        if n_same <= 0:
            a = 0.0  # singleton: no intra-cluster peers
        else:
            a = float(np.mean(dist[i, same & (np.arange(n) != i)]))
        b = float("inf")
        for lab in unique:
            if lab == labels[i]:
                continue
            members = labels == lab
            if not np.any(members):
                continue
            b = min(b, float(np.mean(dist[i, members])))
        if not math.isfinite(b):
            scores.append(0.0)
            continue
        denom = max(a, b)
        scores.append(0.0 if denom == 0.0 else (b - a) / denom)
    return float(np.mean(scores))


def _labels_for_k(dist: np.ndarray, k: int) -> np.ndarray:
    """Cluster labels (1..k) via average linkage on precomputed distances."""
    n = dist.shape[0]
    if k <= 1 or n < 2:
        return np.ones(n, dtype=int)
    condensed = squareform(dist, checks=False)
    Z = linkage(condensed, method="average")
    return fcluster(Z, t=k, criterion="maxclust")


def _choose_k(dist: np.ndarray, k_max: int) -> tuple[int, list[tuple[int, float | str]]]:
    n = dist.shape[0]
    rows: list[tuple[int, float | str]] = []
    best_k, best_s = 1, float("-inf")
    upper = min(k_max, n)
    for k in range(1, upper + 1):
        if k == 1:
            rows.append((1, "N/A"))
            continue
        labels = _labels_for_k(dist, k)
        s = _silhouette_precomputed(dist, labels)
        rows.append((k, s))
        if s > best_s or (s == best_s and k < best_k):
            best_s, best_k = s, k
    # k=1 (silhouette N/A) wins when no positive silhouette exists —
    # e.g. identical samples forced into k=n singletons score 0.
    if best_s <= 0:
        best_k = 1
    return best_k, rows


def _write_dist(
    ws: Worksheet,
    samples: list[str],
    dist: list[list[float | None]],
) -> None:
    ws.append(["", *samples])
    for i, name in enumerate(samples):
        row: list = [name]
        for j in range(len(samples)):
            row.append(dist[i][j])
        ws.append(row)


def _write_suggest(
    ws: Worksheet,
    samples: list[str],
    cluster_ids: list[int | str],
    dist_to_center: list[float | None],
) -> None:
    ws.append(["sample", "cluster_id", "dist_to_center"])
    for name, cid, d in zip(samples, cluster_ids, dist_to_center):
        ws.append([name, cid, d])


def _write_meta(
    ws: Worksheet,
    rows: list[tuple[int, float | str]],
    chosen_k: int,
) -> None:
    ws.append(["k", "silhouette", "chosen"])
    for k, sil in rows:
        ws.append([k, sil, "yes" if k == chosen_k else "no"])


def _write_final(
    ws: Worksheet,
    samples: list[str],
    cluster_ids: list[int | str],
) -> None:
    ws.append(["sample", "cluster_id", "cluster_name"])
    for name, cid in zip(samples, cluster_ids):
        if isinstance(cid, int):
            cname = f"类{cid}"
        else:
            cname = ""
        ws.append([name, cid, cname])


def _verify_cluster_sheets(wb, tag: str, samples: list[str]) -> None:
    """Raise ValueError on L2 check failure.

    Checks: dist symmetric + diag≈0; suggest samples complete;
    meta exactly one chosen=yes; chosen k in range.
    """
    dist_name = f"cluster_dist_{tag}"
    suggest_name = f"cluster_suggest_{tag}"
    meta_name = f"cluster_meta_{tag}"

    for name in (dist_name, suggest_name, meta_name):
        if name not in wb.sheetnames:
            raise ValueError(f"missing sheet {name}")

    n = len(samples)
    ws_d = wb[dist_name]
    # Header row: blank + sample names; body n x n starting at (2,2)
    for i in range(n):
        for j in range(n):
            vij = ws_d.cell(i + 2, j + 2).value
            vji = ws_d.cell(j + 2, i + 2).value
            if i == j:
                if vij is not None and abs(float(vij)) > 1e-9:
                    raise ValueError(f"{dist_name} diag[{i}]={vij!r} not ≈0")
            else:
                if (vij is None) != (vji is None):
                    raise ValueError(
                        f"{dist_name} asymmetry None at ({i},{j}): {vij!r} vs {vji!r}"
                    )
                if vij is not None and vji is not None:
                    if abs(float(vij) - float(vji)) > 1e-9:
                        raise ValueError(
                            f"{dist_name} not symmetric at ({i},{j}): "
                            f"{vij!r} vs {vji!r}"
                        )

    ws_s = wb[suggest_name]
    found: list[str] = []
    for r in range(2, ws_s.max_row + 1):
        v = ws_s.cell(r, 1).value
        if v is not None:
            found.append(str(v).strip())
    expected = [str(s).strip() for s in samples]
    if sorted(found) != sorted(expected):
        raise ValueError(
            f"{suggest_name} samples {found!r} != expected {expected!r}"
        )

    ws_m = wb[meta_name]
    chosen_rows: list[int] = []
    k_values: list[int] = []
    for r in range(2, ws_m.max_row + 1):
        kv = ws_m.cell(r, 1).value
        if kv is not None:
            k_values.append(int(kv))
        if str(ws_m.cell(r, 3).value).strip() == "yes":
            chosen_rows.append(r)
    if len(chosen_rows) != 1:
        raise ValueError(
            f"{meta_name} expected exactly 1 chosen=yes, got {len(chosen_rows)}"
        )
    chosen_k = int(ws_m.cell(chosen_rows[0], 1).value)
    if not k_values:
        raise ValueError(f"{meta_name} has no k rows")
    k_lo, k_hi = min(k_values), max(k_values)
    if not (k_lo <= chosen_k <= k_hi):
        raise ValueError(
            f"{meta_name} chosen k={chosen_k} out of range [{k_lo},{k_hi}]"
        )


def _cluster_angle(
    wb,
    *,
    tag: str,
    freqs: list[float],
    samples: list[str],
    cols: list[list[float | None]],
    f_lo: float,
    f_hi: float,
    k_max: int,
) -> int:
    """Write cluster sheets for one angle. Returns 0, 2, or 3."""
    band_cols = _band_columns(freqs, cols, f_lo, f_hi)
    clusterable = [_is_clusterable(c) for c in band_cols]
    n_clusterable = sum(1 for c in clusterable if c)
    if n_clusterable < 1:
        print(f"no clusterable samples for tag={tag!r}", file=sys.stderr)
        return 2

    n = len(samples)
    dist: list[list[float | None]] = [[None] * n for _ in range(n)]
    for i in range(n):
        for j in range(n):
            if not clusterable[i] or not clusterable[j]:
                dist[i][j] = None
                continue
            if i == j:
                dist[i][j] = 0.0
                continue
            d = _euclid(band_cols[i], band_cols[j])
            dist[i][j] = d

    # Square distance matrix over clusterable samples only
    idx = [i for i in range(n) if clusterable[i]]
    core = np.zeros((n_clusterable, n_clusterable), dtype=float)
    for a, i in enumerate(idx):
        for b, j in enumerate(idx):
            val = dist[i][j]
            core[a, b] = 0.0 if val is None else float(val)

    chosen_k, meta_rows = _choose_k(core, k_max)
    labels_core = _labels_for_k(core, chosen_k)

    # Map labels back; build per-cluster members for centers
    cluster_ids: list[int | str] = ["unclustered"] * n
    members_by_cid: dict[int, list[list[float | None]]] = {}
    for local, global_i in enumerate(idx):
        cid = int(labels_core[local])
        cluster_ids[global_i] = cid
        members_by_cid.setdefault(cid, []).append(band_cols[global_i])

    centers = {cid: _class_center(members) for cid, members in members_by_cid.items()}

    dist_to_center: list[float | None] = []
    for i in range(n):
        cid = cluster_ids[i]
        if not isinstance(cid, int):
            dist_to_center.append(None)
        else:
            dist_to_center.append(_euclid(band_cols[i], centers[cid]))

    dist_name = f"cluster_dist_{tag}"
    suggest_name = f"cluster_suggest_{tag}"
    meta_name = f"cluster_meta_{tag}"
    final_name = f"cluster_final_{tag}"

    _write_dist(_replace_sheet(wb, dist_name), samples, dist)
    _write_suggest(
        _replace_sheet(wb, suggest_name), samples, cluster_ids, dist_to_center
    )
    _write_meta(_replace_sheet(wb, meta_name), meta_rows, chosen_k)

    created_final = False
    if final_name not in wb.sheetnames:
        _write_final(_replace_sheet(wb, final_name), samples, cluster_ids)
        created_final = True

    try:
        _verify_cluster_sheets(wb, tag, samples)
    except ValueError as e:
        for name in (dist_name, suggest_name, meta_name):
            if name in wb.sheetnames:
                del wb[name]
        if created_final and final_name in wb.sheetnames:
            del wb[final_name]
        print(f"VERIFICATION FAIL for {tag!r}: {e}", file=sys.stderr)
        return 3

    return 0


def main(argv: list[str] | None = None) -> int:
    ensure_utf8_stdio()
    parser = argparse.ArgumentParser(
        description="Write cluster_* sheets from delta curves."
    )
    parser.add_argument("--params", required=True, type=Path)
    args = parser.parse_args(argv)

    params = load_params(args.params)
    xlsx_path = resolve_under_output(params, "process_xlsx", args.params)
    angles = params.get("angles") or []
    axial = params.get("axial_angle") or ""
    f_lo = float(params.get("f_lo_hz") or 0)
    f_hi = float(params.get("f_hi_hz") or 0)

    k_max_raw = params.get("cluster_k_max", 5)
    if isinstance(k_max_raw, bool) or not isinstance(k_max_raw, (int, float)):
        print(f"cluster_k_max must be numeric: {k_max_raw!r}", file=sys.stderr)
        return 2
    k_max = int(k_max_raw)
    if k_max <= 0:
        print(f"cluster_k_max must be > 0: {k_max}", file=sys.stderr)
        return 2

    if not angles:
        print("params.angles is empty", file=sys.stderr)
        return 2

    wb = load_workbook(xlsx_path)

    delta_sheets: list[tuple[str, str]] = []
    for angle in angles:
        if angle == axial:
            continue
        tag = normalize_angle_tag(angle)
        name = f"delta_{tag}"
        if name in wb.sheetnames:
            delta_sheets.append((angle, name))

    if not delta_sheets:
        print("no delta_* sheets found; run run_deltas first", file=sys.stderr)
        return 2

    written: list[str] = []
    for angle, delta_name in delta_sheets:
        freqs, samples, cols = read_angle_sheet(wb[delta_name])
        tag = normalize_angle_tag(angle)
        rc = _cluster_angle(
            wb,
            tag=tag,
            freqs=freqs,
            samples=samples,
            cols=cols,
            f_lo=f_lo,
            f_hi=f_hi,
            k_max=k_max,
        )
        if rc != 0:
            if rc == 3:
                wb.save(xlsx_path)
            return rc
        written.append(tag)

    wb.save(xlsx_path)
    print(
        f"VERIFICATION OK: {len(written)} angles, "
        f"cluster_dist/suggest/meta written"
    )
    print(f"run_cluster: wrote sheets for {len(written)} angles: {', '.join(written)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

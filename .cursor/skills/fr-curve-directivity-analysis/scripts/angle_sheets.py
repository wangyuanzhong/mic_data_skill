# -*- coding: utf-8 -*-
"""
Angle-sheet IO and column-wise subtract for directivity analysis.

An "angle sheet" is a wide table in process.xlsx: A column = frequency Hz,
row 1 from column B = sample names, B2.. = amplitude dB for that sample at
that frequency. Each angle (including axial) gets its own sheet, with the
same sample order across sheets so scripts can subtract column-by-column.
"""
from __future__ import annotations

import re
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from quantize import q1


_TAG_SAFE = re.compile(r"[^A-Za-z0-9_]+")


def normalize_angle_tag(label: str) -> str:
    """Normalize an angle label to a safe sheet-name token.

    "90°" -> "90"; "90deg" -> "90"; "off-axis" -> "off_axis"; "30.5" -> "30_5".
    """
    s = str(label).strip()
    # strip trailing degree units
    s = re.sub(r"(?:°|deg(?:rees)?)\s*$", "", s, flags=re.IGNORECASE)
    s = s.strip()
    # replace unsafe runs with single underscore
    s = _TAG_SAFE.sub("_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def angle_sheet_name(label: str) -> str:
    """Sheet name for an angle wide table (= normalize_angle_tag).

    params.angles / axial_angle may keep display labels (e.g. \"0°\");
    process.xlsx sheets always use this normalized token (e.g. \"0\").
    """
    return normalize_angle_tag(label)


def read_angle_sheet(
    ws: Worksheet,
) -> tuple[list[float], list[str], list[list[float | None]]]:
    """Read an angle sheet.

    Returns (freqs, sample_names, columns) where columns[i] is the list of
    values for sample i, aligned with freqs. Blank cells become None.
    """
    samples: list[str] = []
    col = 2
    while True:
        name = ws.cell(1, col).value
        if name is None or str(name).strip() == "":
            break
        samples.append(str(name).strip())
        col += 1
    if not samples:
        raise SystemExit("no sample names on row 1 starting at column B")

    freqs: list[float] = []
    columns: list[list[float | None]] = [[] for _ in samples]
    row = 2
    while True:
        f = ws.cell(row, 1).value
        if f is None:
            break
        freqs.append(float(q1(f)))  # type: ignore[arg-type]
        for i in range(len(samples)):
            raw = ws.cell(row, i + 2).value
            if raw is None or raw == "":
                columns[i].append(None)
            else:
                columns[i].append(q1(raw))
        row += 1
    if not freqs:
        raise SystemExit("no frequency rows in angle sheet")
    return freqs, samples, columns


def assert_columns_aligned(
    *,
    freqs_a: list[float],
    cols_a: list[list[float | None]],
    freqs_b: list[float],
    cols_b: list[list[float | None]],
) -> None:
    """Raise ValueError if two angle sheets are not column-aligned.

    Same frequency grid and same number of sample columns are required so
    that column-wise subtract is meaningful.
    """
    if len(cols_a) != len(cols_b):
        raise ValueError(
            f"column count mismatch: {len(cols_a)} vs {len(cols_b)}"
        )
    if len(freqs_a) != len(freqs_b):
        raise ValueError(
            f"frequency count mismatch: {len(freqs_a)} vs {len(freqs_b)}"
        )
    for i, (fa, fb) in enumerate(zip(freqs_a, freqs_b)):
        if float(fa) != float(fb):
            raise ValueError(
                f"frequency mismatch at row {i}: {fa} vs {fb}"
            )


def subtract_columns(
    a: list[list[float | None]],
    b: list[list[float | None]],
) -> list[list[float | None]]:
    """Column-wise a - b. None on either side -> None on that point.

    Both inputs are `columns` shaped: outer = samples, inner = freqs.
    """
    if len(a) != len(b):
        raise ValueError(
            f"column count mismatch in subtract: {len(a)} vs {len(b)}"
        )
    out: list[list[float | None]] = []
    for col_a, col_b in zip(a, b):
        if len(col_a) != len(col_b):
            raise ValueError("frequency count mismatch within a column")
        col_out: list[float | None] = []
        for va, vb in zip(col_a, col_b):
            if va is None or vb is None:
                col_out.append(None)
            else:
                col_out.append(round(float(va) - float(vb), 6))
        out.append(col_out)
    return out

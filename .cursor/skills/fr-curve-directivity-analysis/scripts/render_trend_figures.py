# -*- coding: utf-8 -*-
"""Render trend (class-mean + per-class overlay) figures from params + xlsx.

Outputs (under <output_dir>/figures/):
  - 走势_类均值_<tag>.png          class-mean curves with selected peak/valley markers
  - 走势_类内叠图_<tag>_类<id>.png  per-class sample delta overlay (+ light class mean)

Requires cluster_final_* and peak_candidates_* (exit 2 if missing).
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
matplotlib.rcParams["font.sans-serif"] = [
    "Microsoft YaHei",
    "SimHei",
    "Noto Sans CJK SC",
    "Arial Unicode MS",
    "DejaVu Sans",
]
matplotlib.rcParams["axes.unicode_minus"] = False
import matplotlib.pyplot as plt  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from openpyxl.worksheet.worksheet import Worksheet  # noqa: E402

from angle_sheets import normalize_angle_tag, read_angle_sheet  # noqa: E402
from params_io import load_params, resolve_under_output  # noqa: E402
from utf8_boot import ensure_utf8_stdio  # noqa: E402


def _prepare_curve(freqs: list[float], vals: list[float | None]):
    xs: list[float] = []
    ys: list[float] = []
    for f, v in zip(freqs, vals):
        if v is None:
            continue
        xs.append(float(f))
        ys.append(float(v))
    return xs, ys


def _setup_log_freq_axis(ax, f_min: float, f_max: float) -> None:
    ax.set_xscale("log")
    ax.set_xlim(f_min, f_max)
    ax.grid(True, which="both", linestyle=":", linewidth=0.4, alpha=0.6)


def _draw_analysis_band(ax, f_lo: float, f_hi: float) -> None:
    ax.axvline(f_lo, color="red", linestyle="--", linewidth=0.8, alpha=0.7)
    ax.axvline(f_hi, color="red", linestyle="--", linewidth=0.8, alpha=0.7)


def _save_fig(fig, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.tight_layout()
    fig.subplots_adjust(bottom=0.18)
    fig.savefig(path, dpi=150)
    plt.close(fig)
    return path


def _read_cluster_final(
    ws: Worksheet,
) -> tuple[list[str], list[int | str], dict[int | str, str]]:
    """Return (samples, cluster_ids, first non-empty name per id)."""
    samples: list[str] = []
    cluster_ids: list[int | str] = []
    names: dict[int | str, str] = {}
    for r in range(2, ws.max_row + 1):
        raw_s = ws.cell(r, 1).value
        if raw_s is None or str(raw_s).strip() == "":
            continue
        sample = str(raw_s).strip()
        raw_id = ws.cell(r, 2).value
        if raw_id is None or str(raw_id).strip() == "":
            continue
        sid = str(raw_id).strip()
        if sid == "unclustered":
            cid: int | str = "unclustered"
        else:
            try:
                cid = int(raw_id)
            except (TypeError, ValueError):
                cid = sid
        raw_name = ws.cell(r, 3).value
        cname = "" if raw_name is None else str(raw_name).strip()
        samples.append(sample)
        cluster_ids.append(cid)
        if cid not in names and cname:
            names[cid] = cname
        elif cid not in names:
            names[cid] = f"类{cid}" if isinstance(cid, int) else str(cid)
    return samples, cluster_ids, names


def _read_selected_peaks(ws: Worksheet) -> list[dict]:
    """Rows from peak_candidates with selected=yes."""
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            continue
        row = {h: ws.cell(r, c).value for c, h in enumerate(headers, start=1)}
        sel = str(row.get("selected") or "").strip().lower()
        if sel != "yes":
            continue
        rows.append(row)
    return rows


def _normalize_cid(raw) -> int | str:
    if raw is None:
        return "unclustered"
    sid = str(raw).strip()
    if sid == "unclustered":
        return "unclustered"
    try:
        return int(raw)
    except (TypeError, ValueError):
        return sid


def _plot_class_means(
    out_path: Path,
    freqs: list[float],
    col_names: list[str],
    columns: list[list[float | None]],
    selected: list[dict],
    name_by_cid: dict[int | str, str],
    angle: str,
    f_lo: float,
    f_hi: float,
) -> Path:
    fig, ax = plt.subplots(figsize=(10, 5.5))
    # map display name -> column index for marker lookup by cluster
    name_to_col: dict[str, list[float | None]] = {
        n: c for n, c in zip(col_names, columns)
    }
    cid_to_name = dict(name_by_cid)

    for name, col in zip(col_names, columns):
        xs, ys = _prepare_curve(freqs, col)
        ax.plot(xs, ys, linewidth=1.4, label=name)

    # selected peak/valley markers
    for row in selected:
        cid = _normalize_cid(row.get("cluster_id"))
        label = cid_to_name.get(cid, f"类{cid}" if isinstance(cid, int) else str(cid))
        f = row.get("freq_hz")
        amp = row.get("amplitude_db")
        if f is None:
            continue
        f = float(f)
        if amp is None:
            # fall back to class-mean curve at nearest freq
            col = name_to_col.get(label)
            if col is None:
                continue
            amp_v = None
            best = float("inf")
            for ff, vv in zip(freqs, col):
                if vv is None:
                    continue
                d = abs(float(ff) - f)
                if d < best:
                    best = d
                    amp_v = float(vv)
            if amp_v is None:
                continue
            amp = amp_v
        else:
            amp = float(amp)
        kind = str(row.get("kind") or "").strip().lower()
        marker = "^" if kind == "peak" else "v"
        color = "C3" if kind == "peak" else "C0"
        ax.scatter(
            [f],
            [amp],
            marker=marker,
            s=48,
            zorder=5,
            color=color,
            edgecolors="k",
            linewidths=0.4,
        )

    if freqs:
        _setup_log_freq_axis(ax, min(freqs) * 0.5, max(freqs) * 2)
    _draw_analysis_band(ax, f_lo, f_hi)
    ax.set_ylabel("角向 − 轴向 (dB)")
    ax.set_title(f"{angle} 类均值走势")
    if len(col_names) <= 12:
        ax.legend(fontsize=7, loc="best")
    return _save_fig(fig, out_path)


def _plot_class_overlay(
    out_path: Path,
    freqs: list[float],
    sample_names: list[str],
    sample_cols: list[list[float | None]],
    mean_freqs: list[float],
    mean_col: list[float | None] | None,
    mean_label: str,
    angle: str,
    cid: int,
    f_lo: float,
    f_hi: float,
) -> Path:
    fig, ax = plt.subplots(figsize=(10, 5.5))
    for name, col in zip(sample_names, sample_cols):
        xs, ys = _prepare_curve(freqs, col)
        ax.plot(xs, ys, linewidth=1.0, alpha=0.85, label=name)
    if mean_col is not None:
        xs, ys = _prepare_curve(mean_freqs, mean_col)
        ax.plot(
            xs,
            ys,
            linewidth=2.0,
            color="k",
            alpha=0.45,
            linestyle="--",
            label=f"{mean_label}(均值)",
        )
    all_fs = list(freqs) + list(mean_freqs)
    if all_fs:
        _setup_log_freq_axis(ax, min(all_fs) * 0.5, max(all_fs) * 2)
    _draw_analysis_band(ax, f_lo, f_hi)
    ax.set_ylabel("角向 − 轴向 (dB)")
    ax.set_title(f"{angle} 类{cid} 类内叠图")
    if len(sample_names) <= 12:
        ax.legend(fontsize=7, loc="best")
    return _save_fig(fig, out_path)


def _render_angle(
    wb,
    tag: str,
    angle: str,
    figures_dir: Path,
    f_lo: float,
    f_hi: float,
) -> list[str]:
    """Render figures for one angle. Caller must ensure final+candidates exist."""
    mean_name = f"class_mean_{tag}"
    delta_name = f"delta_{tag}"
    final_name = f"cluster_final_{tag}"
    cand_name = f"peak_candidates_{tag}"

    written: list[str] = []
    final_samples, final_ids, name_by_cid = _read_cluster_final(wb[final_name])
    selected = _read_selected_peaks(wb[cand_name])

    mean_freqs: list[float] = []
    mean_names: list[str] = []
    mean_cols: list[list[float | None]] = []
    if mean_name in wb.sheetnames:
        mean_freqs, mean_names, mean_cols = read_angle_sheet(wb[mean_name])
        out_mean = figures_dir / f"走势_类均值_{tag}.png"
        _plot_class_means(
            out_mean,
            mean_freqs,
            mean_names,
            mean_cols,
            selected,
            name_by_cid,
            angle,
            f_lo,
            f_hi,
        )
        written.append(out_mean.name)

    if delta_name in wb.sheetnames:
        delta_freqs, delta_samples, delta_cols = read_angle_sheet(wb[delta_name])
        sample_idx = {s: i for i, s in enumerate(delta_samples)}

        members: dict[int, list[tuple[str, list[float | None]]]] = {}
        for sname, cid in zip(final_samples, final_ids):
            if not isinstance(cid, int):
                continue
            if sname not in sample_idx:
                continue
            members.setdefault(cid, []).append(
                (sname, delta_cols[sample_idx[sname]])
            )

        mean_by_name = {n: c for n, c in zip(mean_names, mean_cols)}

        for cid in sorted(members.keys()):
            pairs = members[cid]
            s_names = [p[0] for p in pairs]
            s_cols = [p[1] for p in pairs]
            label = name_by_cid.get(cid, f"类{cid}")
            mean_col = mean_by_name.get(label)
            out_ov = figures_dir / f"走势_类内叠图_{tag}_类{cid}.png"
            _plot_class_overlay(
                out_ov,
                delta_freqs,
                s_names,
                s_cols,
                mean_freqs,
                mean_col,
                label,
                angle,
                cid,
                f_lo,
                f_hi,
            )
            written.append(out_ov.name)

    return written


def main(argv: list[str] | None = None) -> int:
    ensure_utf8_stdio()
    parser = argparse.ArgumentParser(description="Render trend figures.")
    parser.add_argument("--params", required=True, type=Path)
    args = parser.parse_args(argv)

    params = load_params(args.params)
    xlsx_path = resolve_under_output(params, "process_xlsx", args.params)
    output_dir = Path(params.get("output_dir") or args.params.parent)
    figures_dir = output_dir / "figures"
    figures_dir.mkdir(parents=True, exist_ok=True)

    f_lo = float(params.get("f_lo_hz") or 0)
    f_hi = float(params.get("f_hi_hz") or 0)
    angles = params.get("angles") or []
    axial = params.get("axial_angle") or ""

    if not xlsx_path.is_file():
        print(f"process.xlsx not found: {xlsx_path}", file=sys.stderr)
        return 2

    wb = load_workbook(xlsx_path)
    written: list[str] = []

    for angle in angles:
        if angle == axial:
            continue
        tag = normalize_angle_tag(angle)
        final_name = f"cluster_final_{tag}"
        cand_name = f"peak_candidates_{tag}"
        if final_name not in wb.sheetnames or cand_name not in wb.sheetnames:
            print(
                f"missing required sheet for {angle!r}: "
                f"need {final_name} and {cand_name}",
                file=sys.stderr,
            )
            return 2
        written.extend(
            _render_angle(wb, tag, angle, figures_dir, f_lo, f_hi)
        )

    print(f"render_trend_figures: wrote {len(written)} figures: {', '.join(written)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

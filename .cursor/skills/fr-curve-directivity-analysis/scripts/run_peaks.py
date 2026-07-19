# -*- coding: utf-8 -*-
"""
run_peaks.py — class-mean curves + peak/valley candidates.

For each non-axial angle, read cluster_final_<tag> + delta_<tag>, compute
per-cluster mean delta curves, detect peaks (scipy.signal.find_peaks) and
valleys (on -mean), and write class_mean_<tag> + peak_candidates_<tag>.

Usage:
    python run_peaks.py --params <output_dir>/params.json
"""
from __future__ import annotations

import argparse
import math
import sys
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from scipy.signal import find_peaks, peak_widths

from angle_sheets import normalize_angle_tag, read_angle_sheet
from params_io import load_params, resolve_under_output
from utf8_boot import ensure_utf8_stdio

_SELECTED_MATCH_REL = 0.005


def _replace_sheet(wb, name: str) -> Worksheet:
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def _band_mask(freqs: list[float], f_lo: float, f_hi: float) -> list[bool]:
    return [f_lo <= float(f) <= f_hi for f in freqs]


def _band_slice(
    freqs: list[float],
    cols: list[list[float | None]],
    f_lo: float,
    f_hi: float,
) -> tuple[list[float], list[list[float | None]]]:
    mask = _band_mask(freqs, f_lo, f_hi)
    band_freqs = [freqs[i] for i in range(len(freqs)) if mask[i]]
    band_cols = [[col[i] for i in range(len(freqs)) if mask[i]] for col in cols]
    return band_freqs, band_cols


def _read_cluster_final(
    ws: Worksheet,
) -> tuple[list[str], list[int | str], dict[int | str, str]]:
    """Return (samples, cluster_ids, first non-empty name per id)."""
    samples: list[str] = []
    cluster_ids: list[int | str] = []
    names: dict[int | str, str] = {}
    for r in range(2, ws.max_row + 1):
        raw_s = ws.cell(r, 1).value
        if raw_s is None or str(raw_s).strip() == "":
            continue
        sample = str(raw_s).strip()
        raw_id = ws.cell(r, 2).value
        if raw_id is None or str(raw_id).strip() == "":
            continue
        sid = str(raw_id).strip()
        if sid == "unclustered":
            cid: int | str = "unclustered"
        else:
            try:
                cid = int(raw_id)
            except (TypeError, ValueError):
                cid = sid
        raw_name = ws.cell(r, 3).value
        cname = "" if raw_name is None else str(raw_name).strip()
        samples.append(sample)
        cluster_ids.append(cid)
        if cid not in names and cname:
            names[cid] = cname
        elif cid not in names:
            names[cid] = f"类{cid}" if isinstance(cid, int) else str(cid)
    return samples, cluster_ids, names


def _class_mean(
    members: list[list[float | None]],
) -> list[float | None]:
    if not members:
        return []
    n_freq = len(members[0])
    out: list[float | None] = []
    for i in range(n_freq):
        vals = [m[i] for m in members if m[i] is not None]
        if not vals:
            out.append(None)
        else:
            out.append(sum(vals) / len(vals))
    return out


def _to_array(mean: list[float | None]) -> np.ndarray:
    """Finite array for find_peaks; None → very low so it cannot be a peak."""
    y = np.empty(len(mean), dtype=float)
    for i, v in enumerate(mean):
        y[i] = float(v) if v is not None else -1e30
    return y


def _octave_distance(f1: float, f2: float) -> float:
    if f1 <= 0 or f2 <= 0:
        return float("inf")
    return abs(math.log2(f2 / f1))


def _filter_by_octave(
    indices: np.ndarray,
    prominences: np.ndarray,
    freqs: list[float],
    min_octave: float,
) -> list[int]:
    """Keep highest-prominence peaks first; drop those too close in octaves."""
    order = sorted(range(len(indices)), key=lambda i: float(prominences[i]), reverse=True)
    kept: list[int] = []
    kept_freqs: list[float] = []
    for oi in order:
        idx = int(indices[oi])
        f = float(freqs[idx])
        if any(_octave_distance(f, kf) < min_octave for kf in kept_freqs):
            continue
        kept.append(oi)
        kept_freqs.append(f)
    # restore frequency order
    kept.sort(key=lambda oi: int(indices[oi]))
    return kept


def _find_extrema(
    mean: list[float | None],
    freqs: list[float],
    *,
    kind: str,
    prominence_db: float,
    min_octave: float,
) -> list[tuple[int, float, float]]:
    """Return list of (index, amplitude_db, prominence_db) for peak or valley."""
    y = _to_array(mean)
    search = y if kind == "peak" else -y
    indices, props = find_peaks(search, prominence=prominence_db)
    if len(indices) == 0:
        return []
    prominences = np.asarray(props["prominences"], dtype=float)
    kept = _filter_by_octave(indices, prominences, freqs, min_octave)
    out: list[tuple[int, float, float]] = []
    for oi in kept:
        idx = int(indices[oi])
        amp = mean[idx]
        if amp is None:
            continue
        out.append((idx, float(amp), float(prominences[oi])))
    return out


def _interp_freq(freqs: list[float], ip: float) -> float | None:
    """Linear interpolate frequency at fractional sample index."""
    n = len(freqs)
    if n == 0 or not math.isfinite(ip):
        return None
    if ip <= 0:
        return float(freqs[0])
    if ip >= n - 1:
        return float(freqs[n - 1])
    i0 = int(math.floor(ip))
    i1 = i0 + 1
    t = ip - i0
    return float(freqs[i0]) * (1.0 - t) + float(freqs[i1]) * t


def _compute_q(
    mean: list[float | None],
    freqs: list[float],
    idx: int,
    kind: str,
) -> float | str:
    """Q = f0 / Δf_FWHM via half-height width; valleys on -mean. Fail → \"N/A\"."""
    if idx < 0 or idx >= len(freqs) or idx >= len(mean):
        return "N/A"
    f0 = float(freqs[idx])
    if f0 <= 0 or mean[idx] is None:
        return "N/A"
    y = _to_array(mean)
    search = y if kind == "peak" else -y
    try:
        widths, _, left_ips, right_ips = peak_widths(search, np.array([idx]), rel_height=0.5)
    except Exception:
        return "N/A"
    if len(widths) == 0 or not math.isfinite(float(widths[0])) or float(widths[0]) <= 0:
        return "N/A"
    f_left = _interp_freq(freqs, float(left_ips[0]))
    f_right = _interp_freq(freqs, float(right_ips[0]))
    if f_left is None or f_right is None:
        return "N/A"
    df = f_right - f_left
    if df <= 0 or not math.isfinite(df):
        return "N/A"
    q = f0 / df
    if not math.isfinite(q) or q <= 0:
        return "N/A"
    return float(q)


def _normalize_cid(raw) -> int | str:
    if raw is None:
        return ""
    sid = str(raw).strip()
    if sid == "unclustered":
        return "unclustered"
    try:
        return int(raw) if not isinstance(raw, str) else int(sid)
    except (TypeError, ValueError):
        return sid


def _read_selected_yes(ws: Worksheet) -> list[tuple[int | str, str, float]]:
    """Return (cluster_id, kind, freq_hz) for rows with selected=yes."""
    out: list[tuple[int | str, str, float]] = []
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 7).value).strip().lower() != "yes":
            continue
        kind = str(ws.cell(r, 2).value).strip()
        try:
            freq = float(ws.cell(r, 3).value)
        except (TypeError, ValueError):
            continue
        if freq <= 0:
            continue
        out.append((_normalize_cid(ws.cell(r, 1).value), kind, freq))
    return out


def _merge_selected(
    rows: list[tuple],
    old_selected: list[tuple[int | str, str, float]],
) -> list[tuple]:
    """Preserve selected=yes onto nearest new candidate within 0.5% relative freq."""
    if not old_selected or not rows:
        return rows
    # mutable selected flags parallel to rows
    flags = ["no"] * len(rows)
    used: set[int] = set()
    for cid, kind, f_old in old_selected:
        best_i: int | None = None
        best_df = float("inf")
        for i, row in enumerate(rows):
            if i in used:
                continue
            if _normalize_cid(row[0]) != cid or str(row[1]) != kind:
                continue
            f_new = float(row[2])
            if f_old <= 0:
                continue
            if abs(f_new - f_old) / f_old > _SELECTED_MATCH_REL:
                continue
            df = abs(f_new - f_old)
            if df < best_df:
                best_df = df
                best_i = i
        if best_i is not None:
            flags[best_i] = "yes"
            used.add(best_i)
    merged: list[tuple] = []
    for row, sel in zip(rows, flags):
        merged.append((*row[:6], sel))
    return merged


def _write_class_mean(
    ws: Worksheet,
    freqs: list[float],
    col_names: list[str],
    means: list[list[float | None]],
) -> None:
    ws.append(["freq_hz", *col_names])
    for fi, f in enumerate(freqs):
        row: list = [f]
        for mean in means:
            row.append(mean[fi])
        ws.append(row)


def _write_peak_candidates(
    ws: Worksheet,
    rows: list[tuple],
) -> None:
    ws.append(
        [
            "cluster_id",
            "kind",
            "freq_hz",
            "amplitude_db",
            "q",
            "prominence_db",
            "selected",
        ]
    )
    for row in rows:
        ws.append(list(row))


def _peaks_for_angle(
    wb,
    *,
    tag: str,
    freqs: list[float],
    samples: list[str],
    cols: list[list[float | None]],
    f_lo: float,
    f_hi: float,
    prominence_db: float,
    min_octave: float,
) -> int:
    """Write class_mean + peak_candidates for one angle. Returns 0 or 2."""
    final_name = f"cluster_final_{tag}"
    if final_name not in wb.sheetnames:
        print(f"missing {final_name}; run run_cluster first", file=sys.stderr)
        return 2

    band_freqs, band_cols = _band_slice(freqs, cols, f_lo, f_hi)
    final_samples, final_ids, name_by_cid = _read_cluster_final(wb[final_name])

    sample_index = {s: i for i, s in enumerate(samples)}
    members_by_cid: dict[int | str, list[list[float | None]]] = {}
    for sname, cid in zip(final_samples, final_ids):
        if cid == "unclustered":
            continue
        if sname not in sample_index:
            print(
                f"cluster_final sample {sname!r} not in delta_{tag}",
                file=sys.stderr,
            )
            return 2
        members_by_cid.setdefault(cid, []).append(band_cols[sample_index[sname]])

    # Stable order: numeric ids ascending, then other keys
    def _cid_sort_key(cid: int | str):
        if isinstance(cid, int):
            return (0, cid, "")
        return (1, 0, str(cid))

    ordered_cids = sorted(members_by_cid.keys(), key=_cid_sort_key)
    col_names = [name_by_cid.get(cid, f"类{cid}" if isinstance(cid, int) else str(cid))
                 for cid in ordered_cids]
    means = [_class_mean(members_by_cid[cid]) for cid in ordered_cids]

    candidate_rows: list[tuple] = []
    for cid, mean in zip(ordered_cids, means):
        for kind in ("peak", "valley"):
            extrema = _find_extrema(
                mean,
                band_freqs,
                kind=kind,
                prominence_db=prominence_db,
                min_octave=min_octave,
            )
            for idx, amp, prom in extrema:
                q = _compute_q(mean, band_freqs, idx, kind)
                candidate_rows.append(
                    (cid, kind, band_freqs[idx], amp, q, prom, "no")
                )

    mean_name = f"class_mean_{tag}"
    cand_name = f"peak_candidates_{tag}"
    old_selected: list[tuple[int | str, str, float]] = []
    if cand_name in wb.sheetnames:
        old_selected = _read_selected_yes(wb[cand_name])
    candidate_rows = _merge_selected(candidate_rows, old_selected)

    _write_class_mean(_replace_sheet(wb, mean_name), band_freqs, col_names, means)
    _write_peak_candidates(_replace_sheet(wb, cand_name), candidate_rows)
    return 0


def _validate_peak_param(name: str, raw) -> float | None:
    if isinstance(raw, bool) or not isinstance(raw, (int, float)):
        print(f"{name} must be numeric: {raw!r}", file=sys.stderr)
        return None
    val = float(raw)
    if val <= 0:
        print(f"{name} must be > 0: {val}", file=sys.stderr)
        return None
    return val


def main(argv: list[str] | None = None) -> int:
    ensure_utf8_stdio()
    parser = argparse.ArgumentParser(
        description="Write class_mean_* and peak_candidates_* sheets."
    )
    parser.add_argument("--params", required=True, type=Path)
    args = parser.parse_args(argv)

    params = load_params(args.params)
    xlsx_path = resolve_under_output(params, "process_xlsx", args.params)
    angles = params.get("angles") or []
    axial = params.get("axial_angle") or ""
    f_lo = float(params.get("f_lo_hz") or 0)
    f_hi = float(params.get("f_hi_hz") or 0)

    prominence = _validate_peak_param(
        "peak_prominence_db", params.get("peak_prominence_db", 1.0)
    )
    if prominence is None:
        return 2
    min_octave = _validate_peak_param(
        "peak_min_octave", params.get("peak_min_octave", 1.0 / 3.0)
    )
    if min_octave is None:
        return 2

    if not angles:
        print("params.angles is empty", file=sys.stderr)
        return 2

    wb = load_workbook(xlsx_path)

    delta_sheets: list[tuple[str, str]] = []
    for angle in angles:
        if angle == axial:
            continue
        tag = normalize_angle_tag(angle)
        name = f"delta_{tag}"
        if name not in wb.sheetnames:
            print(f"delta sheet {name!r} missing", file=sys.stderr)
            return 2
        delta_sheets.append((angle, name))

    if not delta_sheets:
        print("no non-axial angles to process", file=sys.stderr)
        return 2

    written: list[str] = []
    for angle, delta_name in delta_sheets:
        freqs, samples, cols = read_angle_sheet(wb[delta_name])
        tag = normalize_angle_tag(angle)
        rc = _peaks_for_angle(
            wb,
            tag=tag,
            freqs=freqs,
            samples=samples,
            cols=cols,
            f_lo=f_lo,
            f_hi=f_hi,
            prominence_db=prominence,
            min_octave=min_octave,
        )
        if rc != 0:
            return rc
        written.append(tag)

    wb.save(xlsx_path)
    print(
        f"run_peaks: wrote class_mean/peak_candidates for "
        f"{len(written)} angles: {', '.join(written)}"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())

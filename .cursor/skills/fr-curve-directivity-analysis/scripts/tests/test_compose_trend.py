# -*- coding: utf-8 -*-
"""Tests for compose_html trend (走势分析) section rendering."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from compose_html import compose_report_html
from conftest import build_angle_workbook, write_params
from render_trend_figures import main as render_trend_main
from run_cluster import main as run_cluster_main
from run_deltas import main as run_deltas_main
from run_peaks import main as run_peaks_main


def _prep(tmp_path: Path) -> Path:
    out = tmp_path
    write_params(
        out / "params.json",
        output_dir=out,
        angles=["axis", "90"],
        axial_angle="axis",
        sample_count=2,
        f_lo_hz=100.0,
        f_hi_hz=10000.0,
        extra={
            "cluster_k_max": 5,
            "peak_prominence_db": 0.5,
            "peak_min_octave": 0.1,
            "peak_include_q": True,
            "focus_freqs": [],
        },
    )
    freqs = [100.0, 300.0, 1000.0, 3000.0, 8000.0]
    axis = [[0.0] * 5, [0.0] * 5]
    ang = [[0.0, -0.5, -5.0, -0.5, 0.0], [0.0, -0.5, -5.0, -0.5, 0.0]]
    build_angle_workbook(
        out / "process.xlsx",
        angles=["axis", "90"],
        samples=["S01", "S02"],
        freqs=freqs,
        values_by_angle={"axis": axis, "90": ang},
    )
    run_deltas_main(["--params", str(out / "params.json")])
    run_cluster_main(["--params", str(out / "params.json")])
    run_peaks_main(["--params", str(out / "params.json")])
    # mark at least one selected=yes for peak table
    wb = load_workbook(out / "process.xlsx")
    ws = wb["peak_candidates_90"]
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 2).value) == "peak":
            ws.cell(r, 7).value = "yes"
            break
    wb.save(out / "process.xlsx")
    render_trend_main(["--params", str(out / "params.json")])
    return out


def test_render_trend_section(tmp_path):
    out = _prep(tmp_path)

    html_path = compose_report_html(out)
    text = html_path.read_text(encoding="utf-8")
    assert "走势分析" in text
    assert "S01" in text
    assert "S02" in text
    # peak table headers (Chinese) after selected=yes
    assert "中心频率" in text
    assert "幅度" in text
    assert "Q" in text
    # class-mean figure referenced
    assert "走势_类均值_90" in text
    # section numbering: trend is 5, conclusion is 6
    assert "5. 走势分析" in text
    assert "6. 结论" in text


def test_trend_exit2_when_sheet_missing(tmp_path):
    out = _prep(tmp_path)
    wb = load_workbook(out / "process.xlsx")
    del wb["cluster_final_90"]
    del wb["peak_candidates_90"]
    wb.save(out / "process.xlsx")

    with pytest.raises(SystemExit) as exc:
        compose_report_html(out)
    assert exc.value.code == 2


def test_trend_verification_fail_exit3(tmp_path, monkeypatch):
    out = _prep(tmp_path)
    import compose_html

    orig = compose_html.build_trend_html

    def bad(output_dir, params):
        html = orig(output_dir, params)
        # drop classification table marker to force L3 check fail
        return html.replace("trend-cluster", "X")

    monkeypatch.setattr(compose_html, "build_trend_html", bad)

    with pytest.raises(SystemExit) as exc:
        compose_report_html(out)
    assert exc.value.code == 3
    # L3 fail keeps html for inspection
    assert (out / "report.html").is_file()

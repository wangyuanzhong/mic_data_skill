# -*- coding: utf-8 -*-
from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook
from conftest import build_angle_workbook, write_params
from run_deltas import main as run_deltas_main
from run_cluster import main as run_cluster_main
from run_peaks import main as run_peaks_main


def _prep_clustered(tmp_path: Path):
    out = tmp_path
    write_params(
        out / "params.json",
        output_dir=out,
        angles=["axis", "90"],
        axial_angle="axis",
        sample_count=2,
        f_lo_hz=100.0,
        f_hi_hz=10000.0,
        extra={
            "cluster_k_max": 5,
            "peak_prominence_db": 0.5,
            "peak_min_octave": 0.1,
            "peak_include_q": True,
        },
    )
    # dense-ish freqs for a visible peak at 1000
    freqs = [100.0, 300.0, 1000.0, 3000.0, 8000.0]
    axis = [[0.0] * 5, [0.0] * 5]
    # both samples: peak at 1000 (+5)
    ang = [[0.0, 0.5, 5.0, 0.5, 0.0], [0.0, 0.5, 5.0, 0.5, 0.0]]
    build_angle_workbook(
        out / "process.xlsx",
        angles=["axis", "90"],
        samples=["S01", "S02"],
        freqs=freqs,
        values_by_angle={"axis": axis, "90": ang},
    )
    run_deltas_main(["--params", str(out / "params.json")])
    run_cluster_main(["--params", str(out / "params.json")])
    return out


def test_class_mean_and_peak_candidate(tmp_path):
    out = _prep_clustered(tmp_path)
    rc = run_peaks_main(["--params", str(out / "params.json")])
    assert rc == 0
    wb = load_workbook(out / "process.xlsx")
    assert "class_mean_90" in wb.sheetnames
    assert "peak_candidates_90" in wb.sheetnames
    ws_p = wb["peak_candidates_90"]
    # header: cluster_id, kind, freq_hz, amplitude_db, q, prominence_db, selected
    kinds = [str(ws_p.cell(r, 2).value) for r in range(2, ws_p.max_row + 1)]
    assert "peak" in kinds
    freqs = [float(ws_p.cell(r, 3).value) for r in range(2, ws_p.max_row + 1)
             if str(ws_p.cell(r, 2).value) == "peak"]
    assert any(abs(f - 1000.0) / 1000.0 < 0.05 for f in freqs)
    assert all(str(ws_p.cell(r, 7).value) == "no" for r in range(2, ws_p.max_row + 1))


def test_missing_final_exit2(tmp_path):
    out = _prep_clustered(tmp_path)
    wb = load_workbook(out / "process.xlsx")
    del wb["cluster_final_90"]
    wb.save(out / "process.xlsx")
    assert run_peaks_main(["--params", str(out / "params.json")]) == 2


def test_selected_preserved_on_rerun(tmp_path):
    out = _prep_clustered(tmp_path)
    run_peaks_main(["--params", str(out / "params.json")])
    wb = load_workbook(out / "process.xlsx")
    ws = wb["peak_candidates_90"]
    # select first peak row
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 2).value) == "peak":
            ws.cell(r, 7).value = "yes"
            old_f = float(ws.cell(r, 3).value)
            break
    wb.save(out / "process.xlsx")
    run_peaks_main(["--params", str(out / "params.json")])
    wb = load_workbook(out / "process.xlsx")
    ws = wb["peak_candidates_90"]
    found = False
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 2).value) == "peak" and str(ws.cell(r, 7).value) == "yes":
            assert abs(float(ws.cell(r, 3).value) - old_f) / old_f <= 0.005
            found = True
    assert found


def test_invalid_prominence_exit2(tmp_path):
    out = _prep_clustered(tmp_path)
    import json
    p = json.loads((out / "params.json").read_text(encoding="utf-8"))
    p["peak_prominence_db"] = 0
    (out / "params.json").write_text(json.dumps(p), encoding="utf-8")
    assert run_peaks_main(["--params", str(out / "params.json")]) == 2


def test_peaks_verification_pass(tmp_path, capsys):
    out = _prep_clustered(tmp_path)
    capsys.readouterr()  # drop prep (cluster) VERIFICATION OK
    rc = run_peaks_main(["--params", str(out / "params.json")])
    assert rc == 0
    assert "VERIFICATION OK" in capsys.readouterr().out


def test_peaks_verification_fail_exit3(tmp_path, monkeypatch):
    out = _prep_clustered(tmp_path)
    import run_peaks

    def bad_verify(*a, **k):
        raise ValueError("sabotage")

    monkeypatch.setattr(run_peaks, "_verify_peak_sheets", bad_verify)
    rc = run_peaks_main(["--params", str(out / "params.json")])
    assert rc == 3
    wb = load_workbook(out / "process.xlsx")
    assert "class_mean_90" not in wb.sheetnames
    assert "peak_candidates_90" not in wb.sheetnames

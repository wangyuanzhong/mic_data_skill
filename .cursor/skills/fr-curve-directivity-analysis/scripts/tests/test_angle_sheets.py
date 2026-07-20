# -*- coding: utf-8 -*-
"""Tests for angle_sheets: tag normalize, column align, column subtract."""
from __future__ import annotations

import pytest

from angle_sheets import (
    angle_sheet_name,
    assert_columns_aligned,
    normalize_angle_tag,
    read_angle_sheet,
    subtract_columns,
)


def test_normalize_angle_tag_strips_degree_suffix():
    assert normalize_angle_tag("90°") == "90"
    assert normalize_angle_tag("90deg") == "90"
    assert normalize_angle_tag(" 90 ") == "90"


def test_angle_sheet_name_matches_normalize():
    assert angle_sheet_name("0°") == "0"
    assert angle_sheet_name("90°") == normalize_angle_tag("90°")


def test_normalize_angle_tag_keeps_axis():
    assert normalize_angle_tag("axis") == "axis"


def test_normalize_angle_tag_replaces_unsafe_chars():
    assert normalize_angle_tag("off-axis") == "off_axis"
    assert normalize_angle_tag("30.5") == "30_5"


def test_subtract_columns_aligned():
    a = [[1.0, 2.0], [3.0, 4.0]]
    b = [[0.5, 0.0], [1.0, 1.0]]
    out = subtract_columns(a, b)
    assert out == [[0.5, 2.0], [2.0, 3.0]]


def test_subtract_columns_none_propagates():
    a = [[1.0, None], [None, 4.0]]
    b = [[0.5, 0.0], [1.0, None]]
    out = subtract_columns(a, b)
    assert out == [[0.5, None], [None, None]]


def test_assert_columns_aligned_passes_when_matched():
    assert_columns_aligned(
        freqs_a=[100.0, 200.0],
        cols_a=[[1.0, 2.0], [3.0, 4.0]],
        freqs_b=[100.0, 200.0],
        cols_b=[[5.0, 6.0], [7.0, 8.0]],
    )  # no raise


def test_assert_columns_aligned_raises_on_width_mismatch():
    with pytest.raises(ValueError):
        assert_columns_aligned(
            freqs_a=[100.0],
            cols_a=[[1.0]],
            freqs_b=[100.0],
            cols_b=[[1.0], [2.0]],
        )


def test_assert_columns_aligned_raises_on_freq_mismatch():
    with pytest.raises(ValueError):
        assert_columns_aligned(
            freqs_a=[100.0, 200.0],
            cols_a=[[1.0], [2.0]],
            freqs_b=[100.0, 300.0],
            cols_b=[[1.0], [2.0]],
        )


def test_read_angle_sheet_returns_freqs_samples_columns():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1, "freq_hz")
    ws.cell(1, 2, "S01")
    ws.cell(1, 3, "S02")
    ws.cell(2, 1, 100)
    ws.cell(2, 2, 1.0)
    ws.cell(2, 3, 2.0)
    ws.cell(3, 1, 200)
    ws.cell(3, 2, 3.0)
    ws.cell(3, 3, 4.0)

    freqs, samples, cols = read_angle_sheet(ws)
    assert freqs == [100.0, 200.0]
    assert samples == ["S01", "S02"]
    assert cols == [[1.0, 3.0], [2.0, 4.0]]


def test_read_angle_sheet_treats_blank_as_none():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1, "freq_hz")
    ws.cell(1, 2, "S01")
    ws.cell(2, 1, 100)
    ws.cell(2, 2, 1.0)
    ws.cell(3, 1, 200)
    # leave (3,2) blank

    freqs, samples, cols = read_angle_sheet(ws)
    assert freqs == [100.0, 200.0]
    assert cols == [[1.0, None]]

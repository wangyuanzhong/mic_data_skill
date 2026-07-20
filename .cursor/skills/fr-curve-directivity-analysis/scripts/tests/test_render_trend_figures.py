# -*- coding: utf-8 -*-
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from conftest import build_angle_workbook, write_params
from run_cluster import main as run_cluster_main
from run_deltas import main as run_deltas_main
from run_peaks import main as run_peaks_main
from render_trend_figures import main as render_trend_main


def _prep_peaked(tmp_path: Path) -> Path:
    out = tmp_path
    write_params(
        out / "params.json",
        output_dir=out,
        angles=["axis", "90"],
        axial_angle="axis",
        sample_count=2,
        f_lo_hz=100.0,
        f_hi_hz=10000.0,
        extra={
            "cluster_k_max": 5,
            "peak_prominence_db": 0.5,
            "peak_min_octave": 0.1,
            "peak_include_q": True,
        },
    )
    freqs = [100.0, 300.0, 1000.0, 3000.0, 8000.0]
    axis = [[0.0] * 5, [0.0] * 5]
    ang = [[0.0, -0.5, -5.0, -0.5, 0.0], [0.0, -0.5, -5.0, -0.5, 0.0]]
    build_angle_workbook(
        out / "process.xlsx",
        angles=["axis", "90"],
        samples=["S01", "S02"],
        freqs=freqs,
        values_by_angle={"axis": axis, "90": ang},
    )
    run_deltas_main(["--params", str(out / "params.json")])
    run_cluster_main(["--params", str(out / "params.json")])
    run_peaks_main(["--params", str(out / "params.json")])
    # mark one peak selected so markers path is exercised
    wb = load_workbook(out / "process.xlsx")
    ws = wb["peak_candidates_90"]
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 2).value) == "peak":
            ws.cell(r, 7).value = "yes"
            break
    wb.save(out / "process.xlsx")
    return out


def test_render_trend_figures_writes_pngs(tmp_path):
    out = _prep_peaked(tmp_path)
    rc = render_trend_main(["--params", str(out / "params.json")])
    assert rc == 0
    figures = out / "figures"
    assert (figures / "走势_类均值_90.png").is_file()
    assert (figures / "走势_类内叠图_90_类1.png").is_file()


def test_missing_final_exit2(tmp_path):
    out = _prep_peaked(tmp_path)
    wb = load_workbook(out / "process.xlsx")
    del wb["cluster_final_90"]
    wb.save(out / "process.xlsx")
    assert render_trend_main(["--params", str(out / "params.json")]) == 2


def test_missing_candidates_exit2(tmp_path):
    out = _prep_peaked(tmp_path)
    wb = load_workbook(out / "process.xlsx")
    del wb["peak_candidates_90"]
    wb.save(out / "process.xlsx")
    assert render_trend_main(["--params", str(out / "params.json")]) == 2

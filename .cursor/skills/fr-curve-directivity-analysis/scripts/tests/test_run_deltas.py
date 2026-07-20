# -*- coding: utf-8 -*-
"""Tests for run_deltas: writes delta_<tag> sheet per non-axial angle."""
from __future__ import annotations

import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

from conftest import build_angle_workbook, write_params
from run_deltas import main as run_deltas_main


def _make_batch(tmp_path: Path, *, axis_cols: list[list[float]], angle_cols: list[list[float]]):
    out = tmp_path
    samples = ["S01", "S02"]
    freqs = [100.0, 200.0]
    write_params(
        out / "params.json",
        output_dir=out,
        angles=["axis", "90"],
        axial_angle="axis",
        sample_count=2,
    )
    build_angle_workbook(
        out / "process.xlsx",
        angles=["axis", "90"],
        samples=samples,
        freqs=freqs,
        values_by_angle={"axis": axis_cols, "90": angle_cols},
    )
    return out


def _read_sheet(ws):
    # header row
    samples = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None or str(v).strip() == "":
            break
        samples.append(str(v).strip())
    rows = []
    for r in range(2, ws.max_row + 1):
        f = ws.cell(r, 1).value
        if f is None:
            break
        row = [float(f)]
        for c in range(2, 2 + len(samples)):
            v = ws.cell(r, c).value
            row.append(None if v is None else float(v))
        rows.append(row)
    return samples, rows


def test_run_deltas_writes_delta_sheet_with_constant_offset(tmp_path):
    axis_cols = [[1.0, 3.0], [2.0, 4.0]]
    angle_cols = [[0.0, 2.0], [1.0, 3.0]]  # axis − 1.0 everywhere → delta +1
    out = _make_batch(tmp_path, axis_cols=axis_cols, angle_cols=angle_cols)

    rc = run_deltas_main(["--params", str(out / "params.json")])
    assert rc == 0

    wb = load_workbook(out / "process.xlsx")
    assert "delta_90" in wb.sheetnames
    samples, rows = _read_sheet(wb["delta_90"])
    assert samples == ["S01", "S02"]
    assert rows[0][0] == 100.0
    assert rows[0][1] == 1.0
    assert rows[0][2] == 1.0
    assert rows[1][0] == 200.0
    assert rows[1][1] == 1.0
    assert rows[1][2] == 1.0


def test_run_deltas_skips_axial_sheet(tmp_path):
    axis_cols = [[1.0, 3.0], [2.0, 4.0]]
    angle_cols = [[0.0, 2.0], [1.0, 3.0]]
    out = _make_batch(tmp_path, axis_cols=axis_cols, angle_cols=angle_cols)

    run_deltas_main(["--params", str(out / "params.json")])

    wb = load_workbook(out / "process.xlsx")
    assert "delta_axis" not in wb.sheetnames, "axial angle must not get a delta sheet"


def test_run_deltas_exits_when_column_count_mismatch(tmp_path):
    # angle sheet has 3 sample columns but axis has 2
    out = tmp_path
    write_params(
        out / "params.json",
        output_dir=out,
        angles=["axis", "90"],
        axial_angle="axis",
        sample_count=2,
    )
    build_angle_workbook(
        out / "process.xlsx",
        angles=["axis", "90"],
        samples=["S01", "S02", "S03"],  # 3 samples
        freqs=[100.0, 200.0],
        values_by_angle={
            "axis": [[1.0, 3.0], [2.0, 4.0], [5.0, 6.0]],
            "90": [[2.0, 4.0], [3.0, 5.0], [6.0, 7.0]],
        },
    )
    # but params says sample_count=2; the real check is sheet column alignment
    # so rebuild axis with only 2 cols to force mismatch
    from openpyxl import Workbook

    wb = load_workbook(out / "process.xlsx")
    # delete axis and rewrite with 2 samples
    del wb["axis"]
    ws = wb.create_sheet("axis", 0)
    ws.append(["freq_hz", "S01", "S02"])
    ws.append([100.0, 1.0, 2.0])
    ws.append([200.0, 3.0, 4.0])
    wb.save(out / "process.xlsx")

    rc = run_deltas_main(["--params", str(out / "params.json")])
    assert rc != 0


def test_run_deltas_preserves_existing_angle_sheets(tmp_path):
    axis_cols = [[1.0, 3.0], [2.0, 4.0]]
    angle_cols = [[0.0, 2.0], [1.0, 3.0]]
    out = _make_batch(tmp_path, axis_cols=axis_cols, angle_cols=angle_cols)

    run_deltas_main(["--params", str(out / "params.json")])

    wb = load_workbook(out / "process.xlsx")
    assert "axis" in wb.sheetnames
    assert "90" in wb.sheetnames


def test_run_deltas_looks_up_sheets_by_normalized_tag(tmp_path):
    """params may use '90°' while Agent-written sheets use normalize_angle_tag → '90'."""
    out = tmp_path
    samples = ["S01", "S02"]
    freqs = [100.0, 200.0]
    axis_cols = [[1.0, 3.0], [2.0, 4.0]]
    angle_cols = [[0.0, 2.0], [1.0, 3.0]]  # axis − 1.0 → delta +1
    write_params(
        out / "params.json",
        output_dir=out,
        angles=["0°", "90°"],
        axial_angle="0°",
        sample_count=2,
    )
    build_angle_workbook(
        out / "process.xlsx",
        angles=["0", "90"],  # normalized sheet names (SOP / fill-02)
        samples=samples,
        freqs=freqs,
        values_by_angle={"0": axis_cols, "90": angle_cols},
    )

    rc = run_deltas_main(["--params", str(out / "params.json")])
    assert rc == 0

    wb = load_workbook(out / "process.xlsx")
    assert "delta_90" in wb.sheetnames
    assert "0" in wb.sheetnames and "90" in wb.sheetnames
    samples_out, rows = _read_sheet(wb["delta_90"])
    assert samples_out == samples
    assert rows[0][1] == 1.0

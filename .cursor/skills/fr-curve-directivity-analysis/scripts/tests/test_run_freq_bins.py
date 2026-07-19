# -*- coding: utf-8 -*-
"""Tests for run_freq_bins: 1dB binning of per-angle delta at focus freqs."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from conftest import build_angle_workbook, write_params
from run_deltas import main as run_deltas_main
from run_freq_bins import main as run_freq_bins_main


def _prep(tmp_path: Path, *, axis_cols, angle_cols, focus_freqs):
    out = tmp_path
    write_params(
        out / "params.json",
        output_dir=out,
        angles=["axis", "90"],
        axial_angle="axis",
        sample_count=2,
        extra={"focus_freqs": focus_freqs},
    )
    build_angle_workbook(
        out / "process.xlsx",
        angles=["axis", "90"],
        samples=["S01", "S02"],
        freqs=[100.0, 1000.0, 4000.0],
        values_by_angle={"axis": axis_cols, "90": angle_cols},
    )
    run_deltas_main(["--params", str(out / "params.json")])
    return out


def test_basic_two_freqs(tmp_path):
    axis_cols = [[1.0, 2.0, 3.0], [1.0, 2.0, 3.0]]
    angle_cols = [[2.0, 3.0, 4.0], [4.0, 5.0, 6.0]]  # S01 +1, S02 +3 everywhere
    out = _prep(tmp_path, axis_cols=axis_cols, angle_cols=angle_cols, focus_freqs=[1000, 4000])

    rc = run_freq_bins_main(["--params", str(out / "params.json")])
    assert rc == 0

    wb = load_workbook(out / "process.xlsx")
    assert "freq_bins_90" in wb.sheetnames
    assert "freq_bins_summary_90" in wb.sheetnames

    ws_d = wb["freq_bins_90"]
    assert ws_d.max_column == 1 + 2 * 2  # 样机 + 2 freqs × (差值, 档位)
    assert ws_d.max_row == 1 + 2  # header + 2 samples
    assert str(ws_d.cell(1, 1).value) == "样机"
    assert "1000Hz" in str(ws_d.cell(1, 2).value)
    # S01 delta=+1 -> bin "1~2"; S02 delta=+3 -> bin "3~4"
    assert str(ws_d.cell(2, 3).value) == "1~2"
    assert str(ws_d.cell(3, 3).value) == "3~4"

    ws_s = wb["freq_bins_summary_90"]
    assert ws_s.max_column == 2 * 2
    assert "分组" in str(ws_s.cell(1, 1).value)


def test_none_delta(tmp_path):
    # S02 has a blank cell at 1000Hz in the angle sheet -> delta None at that freq
    axis_cols = [[1.0, 2.0, 3.0], [1.0, 2.0, 3.0]]
    angle_cols = [[2.0, 3.0, 4.0], [2.0, None, 6.0]]  # S02 blank at 1000Hz
    out = _prep(tmp_path, axis_cols=axis_cols, angle_cols=angle_cols, focus_freqs=[1000])

    rc = run_freq_bins_main(["--params", str(out / "params.json")])
    assert rc == 0

    wb = load_workbook(out / "process.xlsx")
    ws_d = wb["freq_bins_90"]
    # S01 delta=+1 -> "1~2"; S02 delta=None -> "N/A"
    assert str(ws_d.cell(2, 3).value) == "1~2"
    assert str(ws_d.cell(3, 3).value) == "N/A"
    # summary: only 1 non-None delta -> total count == 1
    ws_s = wb["freq_bins_summary_90"]
    total = 0
    for r in range(2, ws_s.max_row + 1):
        v = ws_s.cell(r, 2).value
        if v is not None:
            total += int(v)
    assert total == 1


def test_empty_focus_freqs(tmp_path):
    axis_cols = [[1.0, 2.0, 3.0], [1.0, 2.0, 3.0]]
    angle_cols = [[2.0, 3.0, 4.0], [4.0, 5.0, 6.0]]
    out = _prep(tmp_path, axis_cols=axis_cols, angle_cols=angle_cols, focus_freqs=[])

    rc = run_freq_bins_main(["--params", str(out / "params.json")])
    assert rc == 0

    wb = load_workbook(out / "process.xlsx")
    assert "freq_bins_90" not in wb.sheetnames
    assert "freq_bins_summary_90" not in wb.sheetnames


def test_freq_out_of_band(tmp_path, capsys):
    # f_lo=250, f_hi=20000; focus 50Hz is below band
    axis_cols = [[1.0, 2.0, 3.0], [1.0, 2.0, 3.0]]
    angle_cols = [[2.0, 3.0, 4.0], [4.0, 5.0, 6.0]]
    out = _prep(tmp_path, axis_cols=axis_cols, angle_cols=angle_cols, focus_freqs=[50, 1000])

    rc = run_freq_bins_main(["--params", str(out / "params.json")])
    assert rc == 0

    wb = load_workbook(out / "process.xlsx")
    ws_d = wb["freq_bins_90"]
    # only 1000Hz active -> cols = 1 + 1*2 = 3
    assert ws_d.max_column == 3
    assert "1000Hz" in str(ws_d.cell(1, 2).value)
    # 50Hz column must NOT exist
    for c in range(1, ws_d.max_column + 1):
        assert "50Hz" not in str(ws_d.cell(1, c).value)
    captured = capsys.readouterr()
    assert "out of band" in captured.err


def test_all_freqs_out_of_band_exit0(tmp_path, capsys):
    # f_lo=250, f_hi=20000; both focus freqs below band
    axis_cols = [[1.0, 2.0, 3.0], [1.0, 2.0, 3.0]]
    angle_cols = [[2.0, 3.0, 4.0], [4.0, 5.0, 6.0]]
    out = _prep(tmp_path, axis_cols=axis_cols, angle_cols=angle_cols, focus_freqs=[50, 80])

    rc = run_freq_bins_main(["--params", str(out / "params.json")])
    assert rc == 0

    wb = load_workbook(out / "process.xlsx")
    assert "freq_bins_90" not in wb.sheetnames
    captured = capsys.readouterr()
    assert "out of band" in captured.err


def test_idempotent(tmp_path):
    axis_cols = [[1.0, 2.0, 3.0], [1.0, 2.0, 3.0]]
    angle_cols = [[2.0, 3.0, 4.0], [4.0, 5.0, 6.0]]
    out = _prep(tmp_path, axis_cols=axis_cols, angle_cols=angle_cols, focus_freqs=[1000])

    run_freq_bins_main(["--params", str(out / "params.json")])
    run_freq_bins_main(["--params", str(out / "params.json")])

    wb = load_workbook(out / "process.xlsx")
    freq_bins_sheets = [n for n in wb.sheetnames if n.startswith("freq_bins")]
    # exactly 2: freq_bins_90 + freq_bins_summary_90, no duplicates
    assert sorted(freq_bins_sheets) == ["freq_bins_90", "freq_bins_summary_90"]
    ws_d = wb["freq_bins_90"]
    assert ws_d.max_row == 1 + 2  # not doubled


import json


def test_invalid_focus_freqs_exit2(tmp_path):
    axis_cols = [[1.0, 2.0, 3.0], [1.0, 2.0, 3.0]]
    angle_cols = [[2.0, 3.0, 4.0], [4.0, 5.0, 6.0]]
    out = _prep(tmp_path, axis_cols=axis_cols, angle_cols=angle_cols, focus_freqs=[1000])

    bad_values = [[-1000], [0], ["x"], [1000, 1000]]  # negative, zero, non-numeric, dup
    for bv in bad_values:
        (out / "params.json").write_text(
            json.dumps({
                "product": "T", "note": "", "data_root": str(out), "output_dir": str(out),
                "unit": "dB", "f_lo_hz": 250, "f_hi_hz": 20000,
                "process_xlsx": "process.xlsx", "process_md": "process.md",
                "angles": ["axis", "90"], "axial_angle": "axis", "sample_count": 2,
                "envelope": None, "focus_freqs": bv,
            }, ensure_ascii=False),
            encoding="utf-8",
        )
        rc = run_freq_bins_main(["--params", str(out / "params.json")])
        assert rc == 2, f"expected exit 2 for focus_freqs={bv}"


def test_missing_delta_sheet_exit2(tmp_path):
    out = tmp_path
    write_params(
        out / "params.json",
        output_dir=out,
        angles=["axis", "90"],
        axial_angle="axis",
        sample_count=2,
        extra={"focus_freqs": [1000]},
    )
    build_angle_workbook(
        out / "process.xlsx",
        angles=["axis", "90"],
        samples=["S01", "S02"],
        freqs=[100.0, 1000.0],
        values_by_angle={
            "axis": [[1.0, 2.0], [1.0, 2.0]],
            "90": [[2.0, 3.0], [3.0, 4.0]],
        },
    )
    # do NOT run run_deltas -> no delta_90 sheet

    rc = run_freq_bins_main(["--params", str(out / "params.json")])
    assert rc == 2


def test_identical_deltas(tmp_path):
    # all deltas at 1000Hz are exactly +2.0 -> single bin "2~3"
    axis_cols = [[1.0, 2.0, 3.0], [1.0, 2.0, 3.0]]
    angle_cols = [[3.0, 4.0, 5.0], [3.0, 4.0, 5.0]]  # both +2.0
    out = _prep(tmp_path, axis_cols=axis_cols, angle_cols=angle_cols, focus_freqs=[1000])

    rc = run_freq_bins_main(["--params", str(out / "params.json")])
    assert rc == 0

    wb = load_workbook(out / "process.xlsx")
    ws_d = wb["freq_bins_90"]
    assert str(ws_d.cell(2, 3).value) == "2~3"
    assert str(ws_d.cell(3, 3).value) == "2~3"
    ws_s = wb["freq_bins_summary_90"]
    # exactly one bin row, count == 2
    assert ws_s.max_row == 2
    assert str(ws_s.cell(2, 1).value) == "2~3"
    assert int(ws_s.cell(2, 2).value) == 2


def test_verification_pass(tmp_path, capsys):
    axis_cols = [[1.0, 2.0, 3.0], [1.0, 2.0, 3.0]]
    angle_cols = [[2.0, 3.0, 4.0], [4.0, 5.0, 6.0]]
    out = _prep(tmp_path, axis_cols=axis_cols, angle_cols=angle_cols, focus_freqs=[1000, 4000])

    rc = run_freq_bins_main(["--params", str(out / "params.json")])
    assert rc == 0
    captured = capsys.readouterr()
    assert "VERIFICATION OK" in captured.out


def test_verification_fail_exit3(tmp_path, monkeypatch):
    axis_cols = [[1.0, 2.0, 3.0], [1.0, 2.0, 3.0]]
    angle_cols = [[2.0, 3.0, 4.0], [4.0, 5.0, 6.0]]
    out = _prep(tmp_path, axis_cols=axis_cols, angle_cols=angle_cols, focus_freqs=[1000])

    # sabotage: make _bin_label produce an invalid label -> verify must catch -> exit 3
    import run_freq_bins
    monkeypatch.setattr(run_freq_bins, "_bin_label", lambda d: "BAD" if d is not None else "N/A")

    rc = run_freq_bins_main(["--params", str(out / "params.json")])
    assert rc == 3

    wb = load_workbook(out / "process.xlsx")
    assert "freq_bins_90" not in wb.sheetnames, "bad sheet must be deleted on verify fail"
    assert "freq_bins_summary_90" not in wb.sheetnames


def test_verification_fail_deletes_all_angles(tmp_path, monkeypatch):
    # two non-axial angles; sabotage kicks in on first angle -> both must be absent
    out = tmp_path
    write_params(
        out / "params.json",
        output_dir=out,
        angles=["axis", "90", "180"],
        axial_angle="axis",
        sample_count=2,
        extra={"focus_freqs": [1000]},
    )
    build_angle_workbook(
        out / "process.xlsx",
        angles=["axis", "90", "180"],
        samples=["S01", "S02"],
        freqs=[100.0, 1000.0],
        values_by_angle={
            "axis": [[1.0, 2.0], [1.0, 2.0]],
            "90": [[2.0, 3.0], [3.0, 4.0]],
            "180": [[0.0, 1.0], [-1.0, 0.0]],
        },
    )
    run_deltas_main(["--params", str(out / "params.json")])

    import run_freq_bins
    monkeypatch.setattr(run_freq_bins, "_bin_label", lambda d: "BAD" if d is not None else "N/A")

    rc = run_freq_bins_main(["--params", str(out / "params.json")])
    assert rc == 3

    wb = load_workbook(out / "process.xlsx")
    for tag in ("90", "180"):
        assert f"freq_bins_{tag}" not in wb.sheetnames
        assert f"freq_bins_summary_{tag}" not in wb.sheetnames


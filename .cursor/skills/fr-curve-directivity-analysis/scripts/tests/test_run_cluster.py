# -*- coding: utf-8 -*-
"""Tests for run_cluster: Euclidean clustering of delta curves."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from conftest import build_angle_workbook, write_params
from run_deltas import main as run_deltas_main
from run_cluster import cluster_label, main as run_cluster_main


def test_cluster_label_letters():
    assert cluster_label(1) == "CLASS A"
    assert cluster_label(2) == "CLASS B"
    assert cluster_label(26) == "CLASS Z"
    assert cluster_label(27) == "CLASS AA"


def _prep(tmp_path: Path, *, axis_cols, angle_cols, extra=None):
    out = tmp_path
    write_params(
        out / "params.json",
        output_dir=out,
        angles=["axis", "90"],
        axial_angle="axis",
        sample_count=2,
        extra={"focus_freqs": [1000], "cluster_k_max": 5, **(extra or {})},
    )
    build_angle_workbook(
        out / "process.xlsx",
        angles=["axis", "90"],
        samples=["S01", "S02"],
        freqs=[100.0, 1000.0, 4000.0],
        values_by_angle={"axis": axis_cols, "90": angle_cols},
    )
    run_deltas_main(["--params", str(out / "params.json")])
    return out


def test_two_identical_samples_k1(tmp_path):
    # identical deltas → one cluster
    axis = [[1.0, 2.0, 3.0], [1.0, 2.0, 3.0]]
    ang = [[2.0, 3.0, 4.0], [2.0, 3.0, 4.0]]
    out = _prep(tmp_path, axis_cols=axis, angle_cols=ang)

    rc = run_cluster_main(["--params", str(out / "params.json")])
    assert rc == 0

    wb = load_workbook(out / "process.xlsx")
    assert "cluster_dist_90" in wb.sheetnames
    assert "cluster_suggest_90" in wb.sheetnames
    assert "cluster_meta_90" in wb.sheetnames
    assert "cluster_final_90" in wb.sheetnames

    ws_d = wb["cluster_dist_90"]
    assert float(ws_d.cell(2, 2).value) == 0.0  # diagonal
    assert float(ws_d.cell(2, 3).value) == 0.0  # identical

    ws_s = wb["cluster_suggest_90"]
    assert int(ws_s.cell(2, 2).value) == 1
    assert int(ws_s.cell(3, 2).value) == 1
    ws_f = wb["cluster_final_90"]
    assert ws_f.cell(2, 3).value == "CLASS A"
    assert ws_f.cell(3, 3).value == "CLASS A"

    ws_m = wb["cluster_meta_90"]
    # header + rows for k; chosen_k == 1
    chosen = [r for r in range(2, ws_m.max_row + 1) if str(ws_m.cell(r, 3).value) == "yes"]
    assert len(chosen) == 1
    assert int(ws_m.cell(chosen[0], 1).value) == 1


def test_two_clear_groups_k2(tmp_path):
    # Two tight groups of 2 (not all-singletons at k=2): deltas ~+1 vs ~+10
    axis = [[0.0, 0.0, 0.0], [0.0, 0.0, 0.0], [0.0, 0.0, 0.0], [0.0, 0.0, 0.0]]
    ang = [
        [-1.0, -1.0, -1.0],
        [-1.0, -1.0, -1.0],
        [-10.0, -10.0, -10.0],
        [-10.0, -10.0, -10.0],
    ]
    out = tmp_path
    write_params(
        out / "params.json",
        output_dir=out,
        angles=["axis", "90"],
        axial_angle="axis",
        sample_count=4,
        extra={"focus_freqs": [1000], "cluster_k_max": 5},
    )
    build_angle_workbook(
        out / "process.xlsx",
        angles=["axis", "90"],
        samples=["S01", "S02", "S03", "S04"],
        freqs=[100.0, 1000.0, 4000.0],
        values_by_angle={"axis": axis, "90": ang},
    )
    run_deltas_main(["--params", str(out / "params.json")])

    rc = run_cluster_main(["--params", str(out / "params.json")])
    assert rc == 0

    wb = load_workbook(out / "process.xlsx")
    ws_m = wb["cluster_meta_90"]
    chosen = [r for r in range(2, ws_m.max_row + 1) if str(ws_m.cell(r, 3).value) == "yes"]
    assert int(ws_m.cell(chosen[0], 1).value) == 2

    ws_s = wb["cluster_suggest_90"]
    ids = {int(ws_s.cell(r, 2).value) for r in range(2, 6)}
    assert ids == {1, 2}


def test_final_not_overwritten(tmp_path):
    axis = [[1.0, 2.0, 3.0], [1.0, 2.0, 3.0]]
    ang = [[2.0, 3.0, 4.0], [2.0, 3.0, 4.0]]
    out = _prep(tmp_path, axis_cols=axis, angle_cols=ang)
    run_cluster_main(["--params", str(out / "params.json")])
    wb = load_workbook(out / "process.xlsx")
    ws = wb["cluster_final_90"]
    ws.cell(2, 3).value = "自定义类"
    wb.save(out / "process.xlsx")

    run_cluster_main(["--params", str(out / "params.json")])
    wb = load_workbook(out / "process.xlsx")
    assert wb["cluster_final_90"].cell(2, 3).value == "自定义类"


def test_missing_delta_exit2(tmp_path):
    out = tmp_path
    write_params(
        out / "params.json",
        output_dir=out,
        angles=["axis", "90"],
        axial_angle="axis",
        sample_count=2,
        extra={"cluster_k_max": 5},
    )
    build_angle_workbook(
        out / "process.xlsx",
        angles=["axis", "90"],
        samples=["S01", "S02"],
        freqs=[100.0, 1000.0],
        values_by_angle={
            "axis": [[1.0, 2.0], [1.0, 2.0]],
            "90": [[2.0, 3.0], [3.0, 4.0]],
        },
    )
    # no run_deltas
    rc = run_cluster_main(["--params", str(out / "params.json")])
    assert rc == 2


def test_invalid_cluster_k_max_exit2(tmp_path):
    axis = [[1.0, 2.0, 3.0], [1.0, 2.0, 3.0]]
    ang = [[2.0, 3.0, 4.0], [2.0, 3.0, 4.0]]
    out = _prep(tmp_path, axis_cols=axis, angle_cols=ang, extra={"cluster_k_max": 0})
    rc = run_cluster_main(["--params", str(out / "params.json")])
    assert rc == 2


def test_cluster_verification_pass(tmp_path, capsys):
    axis = [[1.0, 2.0, 3.0], [1.0, 2.0, 3.0]]
    ang = [[2.0, 3.0, 4.0], [4.0, 5.0, 6.0]]
    out = _prep(tmp_path, axis_cols=axis, angle_cols=ang)
    rc = run_cluster_main(["--params", str(out / "params.json")])
    assert rc == 0
    assert "VERIFICATION OK" in capsys.readouterr().out


def test_cluster_verification_fail_exit3(tmp_path, monkeypatch):
    axis = [[1.0, 2.0, 3.0], [1.0, 2.0, 3.0]]
    ang = [[2.0, 3.0, 4.0], [4.0, 5.0, 6.0]]
    out = _prep(tmp_path, axis_cols=axis, angle_cols=ang)
    import run_cluster
    def bad_verify(*a, **k):
        raise ValueError("sabotage")
    monkeypatch.setattr(run_cluster, "_verify_cluster_sheets", bad_verify)
    rc = run_cluster_main(["--params", str(out / "params.json")])
    assert rc == 3
    wb = load_workbook(out / "process.xlsx")
    assert "cluster_dist_90" not in wb.sheetnames
    assert "cluster_suggest_90" not in wb.sheetnames
    assert "cluster_meta_90" not in wb.sheetnames


def test_singleton_silhouette_is_zero():
    """Singletons must not inflate mean silhouette (sklearn convention: s=0)."""
    import numpy as np
    from run_cluster import _silhouette_precomputed

    # 3 samples: two identical, one far outlier → labels [1,1,2]
    dist = np.array(
        [
            [0.0, 0.0, 10.0],
            [0.0, 0.0, 10.0],
            [10.0, 10.0, 0.0],
        ]
    )
    labels = np.array([1, 1, 2])
    # If singleton got s=1, mean would be (1+1+1)/3=1.0; correct mean is (1+1+0)/3
    s = _silhouette_precomputed(dist, labels)
    assert abs(s - (1.0 + 1.0 + 0.0) / 3.0) < 1e-9

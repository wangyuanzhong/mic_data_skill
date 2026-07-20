# -*- coding: utf-8 -*-
"""Tests for run_consistency: per-angle cross-sample dispersion of delta curves."""
from __future__ import annotations

import math
from pathlib import Path

from openpyxl import load_workbook

from conftest import build_angle_workbook, write_params
from run_deltas import main as run_deltas_main
from run_consistency import main as run_consistency_main


def _read_summary(ws):
    """Read consistency_summary sheet into list of dicts."""
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, start=1):
            row[h] = ws.cell(r, c).value
        rows.append(row)
    return rows


def _read_consistency_sheet(ws):
    """Read a consistency_<tag> sheet."""
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            break
        row = {}
        for c, h in enumerate(headers, start=1):
            row[h] = ws.cell(r, c).value
        rows.append(row)
    return rows


def _prep(tmp_path: Path, *, axis_cols, angle_cols):
    out = tmp_path
    write_params(
        out / "params.json",
        output_dir=out,
        angles=["axis", "90"],
        axial_angle="axis",
        sample_count=2,
    )
    build_angle_workbook(
        out / "process.xlsx",
        angles=["axis", "90"],
        samples=["S01", "S02"],
        freqs=[100.0, 200.0],
        values_by_angle={"axis": axis_cols, "90": angle_cols},
    )
    run_deltas_main(["--params", str(out / "params.json")])
    return out


def test_consistency_zero_when_samples_agree(tmp_path):
    # both samples have identical delta (axis − 1 everywhere) -> std=0, width=0
    axis_cols = [[1.0, 3.0], [2.0, 4.0]]
    angle_cols = [[0.0, 2.0], [1.0, 3.0]]  # each sample = axis − 1.0
    out = _prep(tmp_path, axis_cols=axis_cols, angle_cols=angle_cols)

    rc = run_consistency_main(["--params", str(out / "params.json")])
    assert rc == 0

    wb = load_workbook(out / "process.xlsx")
    assert "consistency_90" in wb.sheetnames
    rows = _read_consistency_sheet(wb["consistency_90"])
    assert len(rows) == 2
    for row in rows:
        assert row["std_db"] is not None
        assert abs(float(row["std_db"])) < 1e-6
        assert abs(float(row["width_db"])) < 1e-6
        assert row["n_samples"] == 2


def test_consistency_nonzero_when_samples_diverge(tmp_path):
    # sample 1 delta = +1, sample 2 delta = +3 -> std>0, width=2
    axis_cols = [[1.0, 3.0], [2.0, 4.0]]
    angle_cols = [[0.0, 2.0], [-1.0, 1.0]]  # S01 axis−1, S02 axis−3
    out = _prep(tmp_path, axis_cols=axis_cols, angle_cols=angle_cols)

    run_consistency_main(["--params", str(out / "params.json")])

    wb = load_workbook(out / "process.xlsx")
    rows = _read_consistency_sheet(wb["consistency_90"])
    for row in rows:
        assert float(row["width_db"]) == 2.0
        assert float(row["std_db"]) > 0
        assert row["max_db"] is not None
        assert row["min_db"] is not None
        assert float(row["max_db"]) - float(row["min_db"]) == 2.0


def test_consistency_summary_exists(tmp_path):
    axis_cols = [[1.0, 3.0], [2.0, 4.0]]
    angle_cols = [[0.0, 2.0], [1.0, 3.0]]
    out = _prep(tmp_path, axis_cols=axis_cols, angle_cols=angle_cols)

    run_consistency_main(["--params", str(out / "params.json")])

    wb = load_workbook(out / "process.xlsx")
    assert "consistency_summary" in wb.sheetnames
    rows = _read_summary(wb["consistency_summary"])
    assert len(rows) == 1
    assert rows[0]["angle"] == "90"
    assert "max_std_db" in rows[0]
    assert "mean_std_db" in rows[0]
    assert "max_width_db" in rows[0]


def test_consistency_in_band_flag(tmp_path):
    # f_lo=150 -> 100Hz out of band, 200Hz in band
    out = tmp_path
    write_params(
        out / "params.json",
        output_dir=out,
        angles=["axis", "90"],
        axial_angle="axis",
        sample_count=2,
        f_lo_hz=150.0,
        f_hi_hz=20000.0,
    )
    build_angle_workbook(
        out / "process.xlsx",
        angles=["axis", "90"],
        samples=["S01", "S02"],
        freqs=[100.0, 200.0],
        values_by_angle={
            "axis": [[1.0, 3.0], [2.0, 4.0]],
            "90": [[0.0, 2.0], [1.0, 3.0]],
        },
    )
    run_deltas_main(["--params", str(out / "params.json")])

    run_consistency_main(["--params", str(out / "params.json")])

    wb = load_workbook(out / "process.xlsx")
    rows = _read_consistency_sheet(wb["consistency_90"])
    assert rows[0]["in_analysis_band"] == "no"  # 100Hz < 150
    assert rows[1]["in_analysis_band"] == "yes"  # 200Hz in band


def test_consistency_exits_when_delta_missing(tmp_path):
    # no delta sheets exist yet
    out = tmp_path
    write_params(
        out / "params.json",
        output_dir=out,
        angles=["axis", "90"],
        axial_angle="axis",
        sample_count=2,
    )
    build_angle_workbook(
        out / "process.xlsx",
        angles=["axis", "90"],
        samples=["S01", "S02"],
        freqs=[100.0, 200.0],
        values_by_angle={
            "axis": [[1.0, 3.0], [2.0, 4.0]],
            "90": [[2.0, 4.0], [3.0, 5.0]],
        },
    )
    # do NOT run run_deltas first
    rc = run_consistency_main(["--params", str(out / "params.json")])
    assert rc != 0

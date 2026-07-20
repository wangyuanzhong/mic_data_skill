# -*- coding: utf-8 -*-
"""
run_consistency.py — write per-angle cross-sample dispersion of delta curves.

For each `delta_<tag>` sheet, compute per-frequency cross-sample statistics:
  std_db, max_db, min_db, width_db (= max - min), n_samples, in_analysis_band.
Also write a `consistency_summary` sheet with per-angle aggregates over the
analysis band (params.f_lo_hz .. f_hi_hz):
  angle, max_std_db, mean_std_db, max_width_db.

Usage:
    python run_consistency.py --params <output_dir>/params.json
"""
from __future__ import annotations

import argparse
import math
import statistics
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from angle_sheets import normalize_angle_tag, read_angle_sheet
from params_io import load_params, resolve_under_output
from utf8_boot import ensure_utf8_stdio


def _replace_sheet(wb, name: str) -> Worksheet:
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def _stats_for_row(values: list[float | None]) -> dict:
    present = [v for v in values if v is not None]
    n = len(present)
    if n == 0:
        return {
            "std_db": None,
            "max_db": None,
            "min_db": None,
            "width_db": None,
            "n_samples": 0,
        }
    mx = max(present)
    mn = min(present)
    std = statistics.pstdev(present) if n >= 1 else 0.0
    return {
        "std_db": round(std, 6),
        "max_db": round(mx, 6),
        "min_db": round(mn, 6),
        "width_db": round(mx - mn, 6),
        "n_samples": n,
    }


def main(argv: list[str] | None = None) -> int:
    ensure_utf8_stdio()
    parser = argparse.ArgumentParser(description="Write consistency_<tag> sheets.")
    parser.add_argument("--params", required=True, type=Path)
    args = parser.parse_args(argv)

    params = load_params(args.params)
    xlsx_path = resolve_under_output(params, "process_xlsx", args.params)
    angles = params.get("angles") or []
    axial = params.get("axial_angle") or ""
    f_lo = float(params.get("f_lo_hz") or 0)
    f_hi = float(params.get("f_hi_hz") or 0)

    if not angles:
        print("params.angles is empty", file=sys.stderr)
        return 2

    wb = load_workbook(xlsx_path)

    delta_sheets: list[tuple[str, str]] = []  # (angle, sheet_name)
    for angle in angles:
        if angle == axial:
            continue
        tag = normalize_angle_tag(angle)
        name = f"delta_{tag}"
        if name in wb.sheetnames:
            delta_sheets.append((angle, name))

    if not delta_sheets:
        print(
            f"no delta_* sheets found; run run_deltas first",
            file=sys.stderr,
        )
        return 2

    summary_rows: list[list] = []
    written: list[str] = []
    for angle, delta_name in delta_sheets:
        freqs, samples, cols = read_angle_sheet(wb[delta_name])
        tag = normalize_angle_tag(angle)
        sheet_name = f"consistency_{tag}"
        ws = _replace_sheet(wb, sheet_name)
        ws.append(
            ["freq_hz", "std_db", "max_db", "min_db", "width_db", "n_samples", "in_analysis_band"]
        )
        band_stds: list[float] = []
        band_widths: list[float] = []
        for r, f in enumerate(freqs):
            row_vals = [cols[i][r] for i in range(len(samples))]
            stats = _stats_for_row(row_vals)
            in_band = "yes" if (f_lo <= f <= f_hi) else "no"
            ws.append(
                [
                    f,
                    stats["std_db"],
                    stats["max_db"],
                    stats["min_db"],
                    stats["width_db"],
                    stats["n_samples"],
                    in_band,
                ]
            )
            if in_band == "yes" and stats["std_db"] is not None:
                band_stds.append(stats["std_db"])
                band_widths.append(stats["width_db"])
        max_std = max(band_stds) if band_stds else 0.0
        mean_std = sum(band_stds) / len(band_stds) if band_stds else 0.0
        max_width = max(band_widths) if band_widths else 0.0
        summary_rows.append(
            [angle, round(max_std, 6), round(mean_std, 6), round(max_width, 6)]
        )
        written.append(sheet_name)

    ws = _replace_sheet(wb, "consistency_summary")
    ws.append(["angle", "max_std_db", "mean_std_db", "max_width_db"])
    for row in summary_rows:
        ws.append(row)

    wb.save(xlsx_path)
    print(f"run_consistency: wrote {', '.join(written)} + consistency_summary")
    return 0


if __name__ == "__main__":
    sys.exit(main())

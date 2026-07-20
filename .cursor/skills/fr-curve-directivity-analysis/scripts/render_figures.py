# -*- coding: utf-8 -*-
"""Render directivity report figures from params.json + process.xlsx.

Outputs (under <output_dir>/figures/):
  - 差值叠图_<tag>.png  for each non-axial angle (overlay of per-sample delta)
  - 一致性sigma.png     per-angle max_std curve across frequencies

No axial curve figure is produced.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
# Prefer Windows CJK fonts so Chinese titles/legends render.
matplotlib.rcParams["font.sans-serif"] = [
    "Microsoft YaHei",
    "SimHei",
    "Noto Sans CJK SC",
    "Arial Unicode MS",
    "DejaVu Sans",
]
matplotlib.rcParams["axes.unicode_minus"] = False
import matplotlib.pyplot as plt  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from angle_sheets import normalize_angle_tag, read_angle_sheet  # noqa: E402
from params_io import load_params, resolve_under_output  # noqa: E402
from plot_style import (  # noqa: E402
    REPORT_F_MAX_HZ,
    REPORT_F_MIN_HZ,
    draw_analysis_band,
    setup_log_freq_axis,
)
from utf8_boot import ensure_utf8_stdio  # noqa: E402


def _prepare_curve(freqs: list[float], vals: list[float | None]):
    xs: list[float] = []
    ys: list[float] = []
    for f, v in zip(freqs, vals):
        if v is None:
            continue
        xs.append(float(f))
        ys.append(float(v))
    return xs, ys


def _save_fig(fig, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.tight_layout()
    fig.subplots_adjust(bottom=0.18)
    fig.savefig(path, dpi=150)
    plt.close(fig)
    return path


def _plot_delta_overlay(
    out_path: Path,
    freqs: list[float],
    samples: list[str],
    columns: list[list[float | None]],
    angle: str,
    f_lo: float,
    f_hi: float,
) -> Path:
    fig, ax = plt.subplots(figsize=(10, 5.5))
    for name, col in zip(samples, columns):
        xs, ys = _prepare_curve(freqs, col)
        ax.plot(xs, ys, linewidth=1.1, label=name)
    setup_log_freq_axis(ax, REPORT_F_MIN_HZ, REPORT_F_MAX_HZ)
    draw_analysis_band(ax, f_lo, f_hi)
    ax.set_ylabel("轴向 − 角向 (dB)")
    ax.set_title(f"{angle} 相对轴向差值（轴向 − 角向）")
    if len(samples) <= 12:
        ax.legend(fontsize=7, loc="best")
    return _save_fig(fig, out_path)


def _plot_consistency_sigma(
    out_path: Path,
    consistency_rows: list[dict],
    f_lo: float,
    f_hi: float,
) -> Path:
    # group by angle
    by_angle: dict[str, list[tuple[float, float]]] = {}
    for row in consistency_rows:
        angle = row.get("angle")
        f = row.get("freq_hz")
        std = row.get("std_db")
        if angle is None or f is None or std is None:
            continue
        by_angle.setdefault(str(angle), []).append((float(f), float(std)))

    fig, ax = plt.subplots(figsize=(10, 5.5))
    for angle, pts in by_angle.items():
        pts.sort(key=lambda p: p[0])
        xs = [p[0] for p in pts]
        ys = [p[1] for p in pts]
        ax.plot(xs, ys, linewidth=1.2, label=angle)
    setup_log_freq_axis(ax, REPORT_F_MIN_HZ, REPORT_F_MAX_HZ)
    draw_analysis_band(ax, f_lo, f_hi)
    ax.set_ylabel("跨样机 std (dB)")
    ax.set_title("各角度一致性 σ(f)")
    ax.legend(fontsize=8, loc="best")
    return _save_fig(fig, out_path)


def _read_consistency_sheet(ws) -> list[dict]:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            break
        row = {}
        for c, h in enumerate(headers, start=1):
            row[h] = ws.cell(r, c).value
        rows.append(row)
    return rows


def main(argv: list[str] | None = None) -> int:
    ensure_utf8_stdio()
    parser = argparse.ArgumentParser(description="Render directivity figures.")
    parser.add_argument("--params", required=True, type=Path)
    args = parser.parse_args(argv)

    params = load_params(args.params)
    xlsx_path = resolve_under_output(params, "process_xlsx", args.params)
    output_dir = Path(params.get("output_dir") or args.params.parent)
    figures_dir = output_dir / "figures"
    figures_dir.mkdir(parents=True, exist_ok=True)

    f_lo = float(params.get("f_lo_hz") or 0)
    f_hi = float(params.get("f_hi_hz") or 0)
    angles = params.get("angles") or []
    axial = params.get("axial_angle") or ""

    wb = load_workbook(xlsx_path)
    written: list[str] = []

    for angle in angles:
        if angle == axial:
            continue
        tag = normalize_angle_tag(angle)
        sheet_name = f"delta_{tag}"
        if sheet_name not in wb.sheetnames:
            print(f"skip {angle}: {sheet_name} missing", file=sys.stderr)
            continue
        freqs, samples, cols = read_angle_sheet(wb[sheet_name])
        out_path = figures_dir / f"差值叠图_{tag}.png"
        _plot_delta_overlay(out_path, freqs, samples, cols, angle, f_lo, f_hi)
        written.append(out_path.name)

    # consistency sigma (from consistency_<tag> sheets, combined)
    all_rows: list[dict] = []
    for angle in angles:
        if angle == axial:
            continue
        tag = normalize_angle_tag(angle)
        sheet_name = f"consistency_{tag}"
        if sheet_name not in wb.sheetnames:
            continue
        for row in _read_consistency_sheet(wb[sheet_name]):
            row_with_angle = dict(row)
            row_with_angle["angle"] = angle
            all_rows.append(row_with_angle)
    if all_rows:
        out_path = figures_dir / "一致性sigma.png"
        _plot_consistency_sigma(out_path, all_rows, f_lo, f_hi)
        written.append(out_path.name)

    print(f"render_figures: wrote {len(written)} figures: {', '.join(written)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

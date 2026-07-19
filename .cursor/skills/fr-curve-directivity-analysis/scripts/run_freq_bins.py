# -*- coding: utf-8 -*-
"""run_freq_bins.py — write freq_bins_<tag> + freq_bins_summary_<tag> per non-axial angle.

For each non-axial angle's delta_<tag> sheet, for each focus frequency, find the
nearest frequency, take per-sample deltas, bin into 1dB left-closed right-open
bins, and write a detail sheet and a summary sheet.

Usage:
    python run_freq_bins.py --params <output_dir>/params.json
"""
from __future__ import annotations

import argparse
import math
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from angle_sheets import normalize_angle_tag, read_angle_sheet
from params_io import load_params, resolve_under_output
from utf8_boot import ensure_utf8_stdio


_BIN_LABEL_RE = re.compile(r"^-?\d+~-?\d+$")


def _verify_sheets(
    wb,
    angle: str,
    samples: list[str],
    active_freqs: list[float],
) -> None:
    """Raise ValueError on any check failure. Spec §6.1 checks 1-11."""
    tag = normalize_angle_tag(angle)
    detail_name = f"freq_bins_{tag}"
    summary_name = f"freq_bins_summary_{tag}"

    if detail_name not in wb.sheetnames:
        raise ValueError(f"missing sheet {detail_name}")
    if summary_name not in wb.sheetnames:
        raise ValueError(f"missing sheet {summary_name}")

    ws_d = wb[detail_name]
    expected_d_cols = 1 + 2 * len(active_freqs)
    if ws_d.max_column != expected_d_cols:
        raise ValueError(f"{detail_name} cols {ws_d.max_column} != {expected_d_cols}")
    expected_d_rows = 1 + len(samples)
    if ws_d.max_row != expected_d_rows:
        raise ValueError(f"{detail_name} rows {ws_d.max_row} != {expected_d_rows}")

    for si, sname in enumerate(samples):
        v = ws_d.cell(si + 2, 1).value
        if str(v).strip() != str(sname).strip():
            raise ValueError(f"{detail_name} sample R{si+2} mismatch: {v!r} vs {sname!r}")

    for r in range(2, ws_d.max_row + 1):
        for c in range(3, ws_d.max_column + 1, 2):  # bin columns are odd (3,5,7...)
            v = ws_d.cell(r, c).value
            if v is None:
                continue
            s = str(v).strip()
            if s != "N/A" and not _BIN_LABEL_RE.match(s):
                raise ValueError(f"{detail_name} bad bin label R{r}C{c}: {s!r}")
            if s != "N/A":
                lo_s, hi_s = s.split("~")
                if int(lo_s) + 1 != int(hi_s):
                    raise ValueError(f"{detail_name} bad bin width R{r}C{c}: {s!r}")

    ws_s = wb[summary_name]
    expected_s_cols = 2 * len(active_freqs)
    if ws_s.max_column != expected_s_cols:
        raise ValueError(f"{summary_name} cols {ws_s.max_column} != {expected_s_cols}")
    a1 = ws_s.cell(1, 1).value
    if a1 is None or "分组" not in str(a1):
        raise ValueError(f"{summary_name} A1 must be 分组 title, got {a1!r}")

    for fi in range(len(active_freqs)):
        count_col = 2 + fi * 2  # 1-indexed
        total = 0
        for r in range(2, ws_s.max_row + 1):
            v = ws_s.cell(r, count_col).value
            if v is not None:
                total += int(v)
        non_none = sum(
            1 for si in range(len(samples)) if ws_d.cell(si + 2, 2 + fi * 2).value is not None
        )
        if total != non_none:
            raise ValueError(f"{summary_name} freq#{fi} sum {total} != non-None {non_none}")


def _validate_focus_freqs(ff: list) -> list[float] | None:
    if not isinstance(ff, list):
        print("focus_freqs must be a list", file=sys.stderr)
        return None
    seen: set[float] = set()
    out: list[float] = []
    for v in ff:
        if isinstance(v, bool) or not isinstance(v, (int, float)):
            print(f"focus_freqs element not numeric: {v!r}", file=sys.stderr)
            return None
        fv = float(v)
        if fv <= 0:
            print(f"focus_freqs must be > 0: {fv}", file=sys.stderr)
            return None
        if fv in seen:
            print(f"focus_freqs duplicate value: {fv}", file=sys.stderr)
            return None
        seen.add(fv)
        out.append(fv)
    return out


def _format_hz(ff: float) -> str:
    if ff == int(ff):
        return str(int(ff))
    return str(ff)


def _nearest_freq_index(freqs: list[float], target: float) -> int:
    best_i = 0
    best_d = abs(freqs[0] - target)
    for i in range(1, len(freqs)):
        d = abs(freqs[i] - target)
        if d < best_d or (d == best_d and freqs[i] < freqs[best_i]):
            best_i = i
            best_d = d
    return best_i


def _bin_label(d: float | None) -> str:
    if d is None:
        return "N/A"
    lo = math.floor(d)
    return f"{lo}~{lo + 1}"


def _bin_range(present: list[float]) -> tuple[int, int]:
    lo = math.floor(min(present))
    hi = math.ceil(max(present))
    if lo == hi:
        hi = lo + 1
    return lo, hi


def _replace_sheet(wb, name: str) -> Worksheet:
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def main(argv: list[str] | None = None) -> int:
    ensure_utf8_stdio()
    parser = argparse.ArgumentParser(description="Write freq_bins_<tag> sheets.")
    parser.add_argument("--params", required=True, type=Path)
    args = parser.parse_args(argv)

    params = load_params(args.params)
    xlsx_path = resolve_under_output(params, "process_xlsx", args.params)
    angles = params.get("angles") or []
    axial = params.get("axial_angle") or ""
    f_lo = float(params.get("f_lo_hz") or 0)
    f_hi = float(params.get("f_hi_hz") or 0)
    raw_ff = params.get("focus_freqs")
    if raw_ff is None:
        focus_freqs = []
    else:
        validated = _validate_focus_freqs(raw_ff)
        if validated is None:
            return 2
        focus_freqs = validated

    if not angles:
        print("params.angles is empty", file=sys.stderr)
        return 2

    wb = load_workbook(xlsx_path)
    written: list[str] = []

    for angle in angles:
        if angle == axial:
            continue
        tag = normalize_angle_tag(angle)
        delta_name = f"delta_{tag}"
        if delta_name not in wb.sheetnames:
            print(f"delta sheet {delta_name!r} missing", file=sys.stderr)
            return 2
        freqs, samples, cols = read_angle_sheet(wb[delta_name])

        active_freqs: list[float] = []
        per_freq_deltas: list[list[float]] = []
        per_freq_bins: list[list[tuple[int, int]]] = []
        for ff in focus_freqs:
            idx = _nearest_freq_index(freqs, ff)
            nearest = freqs[idx]
            if nearest < f_lo or nearest > f_hi:
                print(f"focus_freq {ff} -> nearest {nearest} out of band, skip", file=sys.stderr)
                continue
            deltas = [cols[si][idx] for si in range(len(samples))]
            present = [d for d in deltas if d is not None]
            if not present:
                lo, hi = 0, 1
            else:
                lo, hi = _bin_range(present)
            counts: dict[int, int] = {b: 0 for b in range(lo, hi)}
            for d in deltas:
                if d is None:
                    continue
                b = max(lo, min(hi - 1, math.floor(d)))
                counts[b] += 1
            active_freqs.append(ff)
            per_freq_deltas.append(deltas)
            per_freq_bins.append([(b, counts[b]) for b in range(lo, hi)])

        if not active_freqs:
            continue

        detail_name = f"freq_bins_{tag}"
        ws_d = _replace_sheet(wb, detail_name)
        headers = ["样机"]
        for ff in active_freqs:
            hz = _format_hz(ff)
            headers.append(f"{hz}Hz差值")
            headers.append(f"{hz}Hz档位")
        ws_d.append(headers)
        for si, sname in enumerate(samples):
            row = [sname]
            for fi in range(len(active_freqs)):
                d = per_freq_deltas[fi][si]
                row.append(d)
                row.append(_bin_label(d))
            ws_d.append(row)

        summary_name = f"freq_bins_summary_{tag}"
        ws_s = _replace_sheet(wb, summary_name)
        sheaders: list[str] = []
        for ff in active_freqs:
            hz = _format_hz(ff)
            sheaders.append(f"{hz}分组")
            sheaders.append(f"{hz}每组数量")
        ws_s.append(sheaders)
        max_rows = max((len(b) for b in per_freq_bins), default=0)
        for r in range(max_rows):
            row: list = []
            for fi in range(len(active_freqs)):
                bins = per_freq_bins[fi]
                if r < len(bins):
                    lo, count = bins[r]
                    row.append(f"{lo}~{lo + 1}")
                    row.append(count)
                else:
                    row.append(None)
                    row.append(None)
            ws_s.append(row)

        try:
            _verify_sheets(wb, angle, samples, active_freqs)
        except ValueError as e:
            # delete ALL freq_bins sheets written so far (current + prior angles)
            # to avoid leaving partial output on multi-angle failure
            if detail_name in wb.sheetnames:
                del wb[detail_name]
            if summary_name in wb.sheetnames:
                del wb[summary_name]
            for name in written:
                if name in wb.sheetnames:
                    del wb[name]
            print(f"VERIFICATION FAIL for {angle!r}: {e}", file=sys.stderr)
            wb.save(xlsx_path)
            return 3

        written.append(detail_name)
        written.append(summary_name)

    wb.save(xlsx_path)
    n_angles = len(written) // 2
    print(
        f"VERIFICATION OK: {n_angles} angles x 2 sheets, "
        f"{len(samples)} samples, {len(active_freqs)} freq points"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())

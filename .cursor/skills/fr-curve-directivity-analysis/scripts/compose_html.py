# -*- coding: utf-8 -*-
"""Compose directivity report.html from figures + Agent notes + sheet tables.

Notes are CLI fill-ins; the script persists them under report_sections/ and
builds the intro table + consistency summary table from params.json /
process.xlsx. The Agent never hand-writes report.html or report_sections/.
"""
from __future__ import annotations

import argparse
import html
import json
import re as _re
import sys
from pathlib import Path

from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import load_workbook

from angle_sheets import normalize_angle_tag
from params_io import resolve_under_output
from utf8_boot import ensure_utf8_stdio, read_text, write_text


def load_section_html(*, inline: str, file_path: str | Path | None) -> str:
    """Prefer file contents when file_path is set; else use inline string."""
    if file_path:
        return read_text(file_path)
    return inline


def _as_note_html(text: str) -> str:
    if not text or not text.strip():
        return ""
    paras = "".join(
        f"<p>{html.escape(line)}</p>" for line in text.splitlines() if line.strip()
    )
    return f'<div class="figure-note">{paras}</div>' if paras else ""


def write_note_files(
    output_dir: Path,
    *,
    intro_note: str = "",
    deltas_note: str = "",
    consistency_note: str = "",
    trend_note: str = "",
    conclusion_note: str = "",
) -> Path:
    sec = Path(output_dir) / "report_sections"
    sec.mkdir(parents=True, exist_ok=True)
    mapping = {
        "intro_note.txt": intro_note,
        "deltas_note.txt": deltas_note,
        "consistency_note.txt": consistency_note,
        "trend_note.txt": trend_note,
        "conclusion_note.txt": conclusion_note,
    }
    for name, text in mapping.items():
        body = text or ""
        if body and not body.endswith("\n"):
            body += "\n"
        write_text(sec / name, body)
    return sec


def build_intro_html(params: dict, intro_note: str) -> str:
    rows = [
        ("产品", params.get("product") or ""),
        ("说明", params.get("note") or "无"),
        ("幅度单位", params.get("unit") or ""),
        ("分析频段", f"{params.get('f_lo_hz')}–{params.get('f_hi_hz')} Hz"),
        ("角度列表", ", ".join(params.get("angles") or []) or "—"),
        ("轴向参考", params.get("axial_angle") or "—"),
        ("样机数", str(params.get("sample_count") or "—")),
    ]
    body = "<table><tbody>"
    for k, v in rows:
        body += f"<tr><th>{html.escape(k)}</th><td>{html.escape(v)}</td></tr>"
    body += "</tbody></table>"
    if intro_note.strip():
        body += _as_note_html(intro_note)
    return body


def build_consistency_summary_html(output_dir: Path) -> str:
    xlsx_path = output_dir / "process.xlsx"
    if not xlsx_path.is_file():
        return ""
    wb = load_workbook(xlsx_path)
    if "consistency_summary" not in wb.sheetnames:
        return ""
    ws = wb["consistency_summary"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    body = '<h3>各角度一致性汇总</h3><table><thead><tr>'
    body += "".join(f"<th>{html.escape(str(h or ''))}</th>" for h in headers)
    body += "</tr></thead><tbody>"
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            break
        body += "<tr>"
        for c in range(1, len(headers) + 1):
            v = ws.cell(r, c).value
            body += f"<td>{html.escape(str(v if v is not None else ''))}</td>"
        body += "</tr>"
    body += "</tbody></table>"
    return body


def build_freq_bins_html(output_dir: Path, params: dict) -> str:
    """Render freq_bins detail + summary tables per non-axial angle.

    Returns "" if focus_freqs is empty or no freq_bins sheets exist.
    Raises SystemExit(2) if focus_freqs non-empty but a required sheet is missing.
    """
    focus_freqs = params.get("focus_freqs") or []
    if not focus_freqs:
        return ""
    angles = params.get("angles") or []
    axial = params.get("axial_angle") or ""
    xlsx_path = output_dir / "process.xlsx"
    if not xlsx_path.is_file():
        return ""
    wb = load_workbook(xlsx_path)

    body = ''
    for angle in angles:
        if angle == axial:
            continue
        tag = normalize_angle_tag(angle)
        detail_name = f"freq_bins_{tag}"
        summary_name = f"freq_bins_summary_{tag}"
        if detail_name not in wb.sheetnames:
            raise SystemExit(2)
        if summary_name not in wb.sheetnames:
            raise SystemExit(2)
        body += f'<h3>角度 {html.escape(str(angle))}</h3>'

        # detail table
        ws_d = wb[detail_name]
        body += '<table class="freq-bins-detail"><thead><tr>'
        for c in range(1, ws_d.max_column + 1):
            h = ws_d.cell(1, c).value
            body += f"<th>{html.escape(str(h or ''))}</th>"
        body += "</tr></thead><tbody>"
        for r in range(2, ws_d.max_row + 1):
            body += "<tr>"
            for c in range(1, ws_d.max_column + 1):
                v = ws_d.cell(r, c).value
                body += f"<td>{html.escape(str(v if v is not None else ''))}</td>"
            body += "</tr>"
        body += "</tbody></table>"

        # summary table
        ws_s = wb[summary_name]
        body += '<table class="freq-bins-summary"><thead><tr>'
        for c in range(1, ws_s.max_column + 1):
            h = ws_s.cell(1, c).value
            body += f"<th>{html.escape(str(h or ''))}</th>"
        body += "</tr></thead><tbody>"
        for r in range(2, ws_s.max_row + 1):
            if all(ws_s.cell(r, c).value is None for c in range(1, ws_s.max_column + 1)):
                continue
            body += "<tr>"
            for c in range(1, ws_s.max_column + 1):
                v = ws_s.cell(r, c).value
                body += f"<td>{html.escape(str(v if v is not None else ''))}</td>"
            body += "</tr>"
        body += "</tbody></table>"
    return body


def _figure_block(src: str, caption: str) -> str:
    return (
        '<div class="figure">'
        f'<img src="{html.escape(src)}" alt="{html.escape(caption)}" />'
        f'<div class="caption">{html.escape(caption)}</div>'
        "</div>"
    )


def build_trend_html(output_dir: Path, params: dict) -> str:
    """Render 走势分析 body (no h2): cluster table, class-mean fig, peaks, overlays.

    Raises SystemExit(2) if a required cluster_final_* or peak_candidates_* sheet
    is missing for any non-axial angle.
    """
    angles = params.get("angles") or []
    axial = params.get("axial_angle") or ""
    xlsx_path = Path(output_dir) / "process.xlsx"
    if not xlsx_path.is_file():
        raise SystemExit(2)
    wb = load_workbook(xlsx_path)
    figures_dir = Path(output_dir) / "figures"

    body = ""
    for angle in angles:
        if angle == axial:
            continue
        tag = normalize_angle_tag(angle)
        final_name = f"cluster_final_{tag}"
        cand_name = f"peak_candidates_{tag}"
        if final_name not in wb.sheetnames or cand_name not in wb.sheetnames:
            raise SystemExit(2)

        body += f'<h3>角度 {html.escape(str(angle))}</h3>'

        # classification table from cluster_final
        ws_f = wb[final_name]
        body += '<table class="trend-cluster"><thead><tr>'
        body += "".join(
            f"<th>{html.escape(h)}</th>" for h in ("样机", "类", "类名")
        )
        body += "</tr></thead><tbody>"
        for r in range(2, ws_f.max_row + 1):
            sample = ws_f.cell(r, 1).value
            if sample is None or str(sample).strip() == "":
                continue
            cid = ws_f.cell(r, 2).value
            cname = ws_f.cell(r, 3).value
            body += (
                "<tr>"
                f"<td>{html.escape(str(sample))}</td>"
                f"<td>{html.escape(str(cid if cid is not None else ''))}</td>"
                f"<td>{html.escape(str(cname if cname is not None else ''))}</td>"
                "</tr>"
            )
        body += "</tbody></table>"

        # class-mean figure
        mean_name = f"走势_类均值_{tag}.png"
        if (figures_dir / mean_name).is_file():
            src = (Path("figures") / mean_name).as_posix()
            body += _figure_block(src, Path(mean_name).stem)

        # peak table: selected=yes only
        ws_p = wb[cand_name]
        headers = [ws_p.cell(1, c).value for c in range(1, ws_p.max_column + 1)]
        try:
            sel_idx = headers.index("selected") + 1
        except ValueError:
            sel_idx = 7
        selected_rows: list[list] = []
        for r in range(2, ws_p.max_row + 1):
            if ws_p.cell(r, 1).value is None:
                continue
            sel = str(ws_p.cell(r, sel_idx).value or "").strip().lower()
            if sel != "yes":
                continue
            selected_rows.append(
                [
                    ws_p.cell(r, 1).value,  # cluster_id
                    ws_p.cell(r, 2).value,  # kind
                    ws_p.cell(r, 3).value,  # freq_hz
                    ws_p.cell(r, 4).value,  # amplitude_db
                    ws_p.cell(r, 5).value,  # q
                ]
            )
        if selected_rows:
            peak_headers = ("类", "类型", "中心频率", "幅度", "Q")
            body += '<table class="trend-peaks"><thead><tr>'
            body += "".join(f"<th>{html.escape(h)}</th>" for h in peak_headers)
            body += "</tr></thead><tbody>"
            for row in selected_rows:
                body += "<tr>"
                for v in row:
                    body += f"<td>{html.escape(str(v if v is not None else ''))}</td>"
                body += "</tr>"
            body += "</tbody></table>"

        # class overlay figures
        overlays = sorted(figures_dir.glob(f"走势_类内叠图_{tag}_类*.png"))
        for p in overlays:
            src = (Path("figures") / p.name).as_posix()
            body += _figure_block(src, p.stem)

    return body


def _verify_trend_html(text: str, params: dict) -> None:
    """Spec §6.3 L3 trend checks. Raise ValueError on failure.

    Always runs (SP3 default always produces 走势分析), even when focus_freqs is empty.
    """
    if "走势分析" not in text:
        raise ValueError("missing 走势分析 section title")
    angles = params.get("angles") or []
    axial = params.get("axial_angle") or ""
    non_axial = [a for a in angles if a != axial]
    if not non_axial:
        return
    n_cluster = len(_re.findall(r'class="trend-cluster"', text))
    if n_cluster != len(non_axial):
        raise ValueError(
            f"trend cluster tables {n_cluster} != non-axial angles {len(non_axial)}"
        )
    # If selected-peak tables were rendered, require peak/valley column headers
    if 'class="trend-peaks"' in text:
        for col in ("中心频率", "幅度", "Q"):
            if col not in text:
                raise ValueError(f"missing peak table column {col}")


def _verify_report_html(html_path: Path, params: dict) -> None:
    """Spec §6.2 freq_bins L3 + §6.3 trend L3. Raise ValueError on failure."""
    text = html_path.read_text(encoding="utf-8")
    _verify_trend_html(text, params)

    focus_freqs = params.get("focus_freqs") or []
    if not focus_freqs:
        return  # freq_bins section skipped
    if "频点差值分档" not in text:
        raise ValueError("missing 频点差值分档 section title")
    angles = params.get("angles") or []
    axial = params.get("axial_angle") or ""
    non_axial = [a for a in angles if a != axial]
    if not non_axial:
        return

    n_detail = len(_re.findall(r'class="freq-bins-detail"', text))
    n_summary = len(_re.findall(r'class="freq-bins-summary"', text))
    if n_detail != len(non_axial):
        raise ValueError(f"detail tables {n_detail} != non-axial angles {len(non_axial)}")
    if n_summary != len(non_axial):
        raise ValueError(f"summary tables {n_summary} != non-axial angles {len(non_axial)}")
    if "每组数量" not in text:
        raise ValueError("missing 每组数量 header in summary table")

    expected_d_cols = 1 + 2 * len(focus_freqs)
    expected_s_cols = 2 * len(focus_freqs)

    m = _re.search(r'class="freq-bins-detail"><thead><tr>(.*?)</tr>', text)
    if not m:
        raise ValueError("cannot parse detail table header")
    d_th_count = len(_re.findall(r"<th>", m.group(1)))
    if d_th_count != expected_d_cols:
        raise ValueError(f"detail table cols {d_th_count} != {expected_d_cols}")

    m = _re.search(r'class="freq-bins-summary"><thead><tr>(.*?)</tr>', text)
    if not m:
        raise ValueError("cannot parse summary table header")
    s_th_count = len(_re.findall(r"<th>", m.group(1)))
    if s_th_count != expected_s_cols:
        raise ValueError(f"summary table cols {s_th_count} != {expected_s_cols}")


def list_delta_figures(figures_dir: Path) -> list[Path]:
    return sorted(figures_dir.glob("差值叠图_*.png"))


def list_consistency_figures(figures_dir: Path) -> list[Path]:
    return sorted(figures_dir.glob("一致性sigma*.png"))


def compose_report_html(
    output_dir: Path,
    *,
    intro_note: str = "",
    deltas_note: str = "",
    consistency_note: str = "",
    trend_note: str = "",
    conclusion_note: str = "",
    figures_dir: Path | None = None,
    persist_notes: bool = True,
) -> Path:
    output_dir = Path(output_dir)
    if persist_notes:
        write_note_files(
            output_dir,
            intro_note=intro_note,
            deltas_note=deltas_note,
            consistency_note=consistency_note,
            trend_note=trend_note,
            conclusion_note=conclusion_note,
        )
    figures_dir = Path(figures_dir) if figures_dir else output_dir / "figures"
    params: dict = {}
    params_path = output_dir / "params.json"
    if params_path.is_file():
        params = json.loads(read_text(params_path))

    delta_blocks = [
        {"src": (Path("figures") / p.name).as_posix(), "caption": p.stem}
        for p in list_delta_figures(figures_dir)
    ]
    consistency_blocks = [
        {"src": (Path("figures") / p.name).as_posix(), "caption": p.stem}
        for p in list_consistency_figures(figures_dir)
    ]

    intro_html = build_intro_html(params, intro_note)
    deltas_notes_html = _as_note_html(deltas_note)
    consistency_notes_html = _as_note_html(consistency_note)
    freq_bins_html = build_freq_bins_html(output_dir, params)
    consistency_summary_html = build_consistency_summary_html(output_dir)
    trend_notes_html = _as_note_html(trend_note)
    trend_html = build_trend_html(output_dir, params)
    conclusion_html = _as_note_html(conclusion_note) or "<p></p>"

    tpl_dir = Path(__file__).resolve().parent.parent / "assets" / "templates"
    env = Environment(
        loader=FileSystemLoader(str(tpl_dir)),
        autoescape=select_autoescape(["html", "xml"]),
    )
    css = read_text(tpl_dir / "report.css")
    tpl = env.get_template("report.html.j2")
    rendered = tpl.render(
        product=params.get("product") or "指向性分析报告",
        css=css,
        intro_html=intro_html,
        deltas_notes_html=deltas_notes_html,
        delta_figures=delta_blocks,
        freq_bins_html=freq_bins_html,
        consistency_notes_html=consistency_notes_html,
        consistency_summary_html=consistency_summary_html,
        consistency_figures=consistency_blocks,
        trend_notes_html=trend_notes_html,
        trend_html=trend_html,
        conclusion_html=conclusion_html,
    )
    out = output_dir / "report.html"
    write_text(out, rendered)
    try:
        _verify_report_html(out, params)
    except ValueError as e:
        print(f"REPORT VERIFICATION FAIL: {e}", file=sys.stderr)
        raise SystemExit(3)
    return out


def main() -> None:
    ensure_utf8_stdio()
    p = argparse.ArgumentParser(
        description="Compose directivity report.html (script tables + Agent notes)"
    )
    p.add_argument("--output-dir", required=True)
    p.add_argument("--intro-note", default="")
    p.add_argument("--intro-note-file", default=None)
    p.add_argument("--deltas-note", default="")
    p.add_argument("--deltas-note-file", default=None)
    p.add_argument("--consistency-note", default="")
    p.add_argument("--consistency-note-file", default=None)
    p.add_argument("--trend-note", default="")
    p.add_argument("--trend-note-file", default=None)
    p.add_argument("--conclusion-note", default="")
    p.add_argument("--conclusion-note-file", default=None)
    args = p.parse_args()

    path = compose_report_html(
        Path(args.output_dir),
        intro_note=load_section_html(inline=args.intro_note, file_path=args.intro_note_file),
        deltas_note=load_section_html(inline=args.deltas_note, file_path=args.deltas_note_file),
        consistency_note=load_section_html(
            inline=args.consistency_note, file_path=args.consistency_note_file
        ),
        trend_note=load_section_html(inline=args.trend_note, file_path=args.trend_note_file),
        conclusion_note=load_section_html(
            inline=args.conclusion_note, file_path=args.conclusion_note_file
        ),
    )
    print(f"OK wrote {path}")


if __name__ == "__main__":
    main()
